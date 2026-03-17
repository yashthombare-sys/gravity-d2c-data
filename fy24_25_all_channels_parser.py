#!/usr/bin/env python3
"""
Parse FY24-FY25 MIS.xlsx → extract ALL channels (Shopify, Amazon, Flipkart, FirstCry, Blinkit, Instamart)
and generate JSON data files + update dashboard.html with the data.
"""
import openpyxl
import json
import os
import re

XLSX = 'FY24-FY25 MIS.xlsx'

SHEET_MAP = {
    'April24': 'Apr 2024', 'May': 'May 2024', 'June': 'Jun 2024',
    'July': 'Jul 2024', 'Aug': 'Aug 2024', 'Sept': 'Sep 2024',
    'Oct': 'Oct 2024', 'Nov': 'Nov 2024', 'Dec': 'Dec 2024',
    'Jan': 'Jan 2025', 'Feb 25': 'Feb 2025', 'March 25': 'Mar 2025'
}

# Channel name normalization
CHANNEL_NORM = {
    'Shopify': 'D2C', 'Amazon': 'Amazon', 'Flipkart': 'Flipkart',
    'Firstcry': 'FirstCry', 'FirstCry': 'FirstCry',
    'Blinkit': 'Blinkit', 'Instamart': 'Instamart',
}

COGS_MAP = {
    "V1": 225, "V2": 275, "V3": 662, "V4": 170, "V5": 225, "V6": 275,
    "V7": 600, "V7 Police Cruiser": 600, "V8": 700, "V9": 778, "V10": 1009,
    "V1- P of 2": 531, "V2- P of 2": 649, "V4- P of 2": 401, "V4- P of 3": 368,
    "V6- P of 2": 649, "V6- P of 3": 924, "V9 P of 2": 1664,
    "Busy Book Pink": 300, "Busy Book Blue": 300, "Busy Book": 300, "Human Book": 300,
    "V9-V3 Combo": 1440, "V9-V10 Combo": 1787, "V1-V4 Combo": 404,
    "V6-V2 Combo": 612, "V6-V4 Combo": 445, "V1-V2 Combo": 524,
    "V2-V4 Combo": 488, "V9-V2 Combo": 488, "V6-V1 Combo": 608,
    "V7-V4 Combo": 770, "V4-V5 Combo": 395,
    "Ganesha": 290, "Krishna": 290, "Hanuman": 290,
    "Car": 540, "Tank": 862, "JCB": 540,
    "CS Basics 1": 250, "Drawing Board": 250,
    "BB": 300,  # Amazon/Flipkart shorthand for Busy Book
    "V1V4": 404, "V1V2": 524, "V1V6": 500, "V2V4": 488, "V2V6": 612, "V4V2": 488,
    "V6V2": 612,
}

# SKU normalization for Amazon/Flipkart shorthand names
NAME_NORM = {
    'BB': 'Busy Book Blue',
    'V1V4': 'V1-V4 Combo', 'V1V2': 'V1-V2 Combo', 'V1V6': 'V1-V6 Combo',
    'V2V4': 'V2-V4 Combo', 'V2V6': 'V6-V2 Combo', 'V4V2': 'V2-V4 Combo',
    'V6V2': 'V6-V2 Combo',
    'Busy Book': 'Busy Book Blue',
}

SKIP_NAMES = {'Total', 'Total Revenue', 'Total Revenue ', 'Total Expance',
              'Net P&L', 'P&L', 'P&L %', 'COGS', 'Logistics', 'Marketing',
              'Shopify', 'Grand Total', 'TOTAL', 'Blended', 'AOV', 'CAC',
              'Amazon Ex', 'Flipkart Ex', 'Logitics', 'Ad Spent',
              'Commisions', 'Amazon', 'Flipkart', 'Blinkit', 'Firstcry',
              'FirstCry', 'Instamart', 'Ad spent'}

def num(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return v
    if isinstance(v, str):
        v = v.strip().replace(',', '').replace('₹', '')
        if v in ('', '-', '#DIV/0!', '#REF!', '#VALUE!'): return 0
        try: return float(v)
        except: return 0
    return 0


def find_sections(ws):
    """Find row ranges for each channel section in a worksheet."""
    sections = []
    for r in range(1, ws.max_row + 1):
        for c in (1, 2):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and v.strip() in CHANNEL_NORM:
                sections.append((r, CHANNEL_NORM[v.strip()]))
    # Add end boundary
    result = []
    for i, (start_row, channel) in enumerate(sections):
        end_row = sections[i+1][0] - 1 if i + 1 < len(sections) else ws.max_row
        result.append((channel, start_row, end_row))
    return result


def find_header_row(ws, start_row, end_row):
    """Find the header row (contains 'Products' or starts with ' ') within a range."""
    for r in range(start_row, min(start_row + 10, end_row + 1)):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str) and v.strip() in ('Products', ''):
            return r
    return None


def parse_d2c_section(ws, start_row, end_row):
    """Parse Shopify/D2C section."""
    header_row = find_header_row(ws, start_row, end_row)
    if not header_row:
        return {}, 0

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h and isinstance(h, str):
            headers[h.strip()] = c

    col_rev = headers.get('Total revenue', headers.get('Total Revenue', None))
    col_orders = headers.get('Total orders', None)
    col_shipped = headers.get('Shipped', None)
    col_delivered = headers.get('Delivered', None)
    col_rto = headers.get('RTO', None)
    col_freight = headers.get('Shipping charges', None)
    col_adspent = headers.get('Ad spent', headers.get('Ad Spent', None))
    col_intransit = headers.get('In-transit', headers.get('In-T Count', None))

    # Get total ad spend from Marketing summary row
    total_ad = 0
    for r in range(header_row, end_row + 1):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str) and v.strip() == 'Marketing':
            total_ad = num(ws.cell(r, 2).value)
            break

    # If no Marketing row, sum product-level ad spent
    if total_ad == 0 and col_adspent:
        for r in range(header_row + 1, end_row + 1):
            name = ws.cell(r, 1).value
            if not name or not isinstance(name, str): continue
            if name.strip() in SKIP_NAMES or name.strip().startswith('Total'): break
            total_ad += num(ws.cell(r, col_adspent).value)

    products = {}
    for r in range(header_row + 1, end_row + 1):
        name = ws.cell(r, 1).value
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if name in SKIP_NAMES or name.startswith('Total'): break
        name = NAME_NORM.get(name, name)

        revenue = num(ws.cell(r, col_rev).value) if col_rev else 0
        orders = int(num(ws.cell(r, col_orders).value)) if col_orders else 0
        if revenue == 0 and orders == 0: continue

        shipped = int(num(ws.cell(r, col_shipped).value)) if col_shipped else 0
        delivered = int(num(ws.cell(r, col_delivered).value)) if col_delivered else 0
        rto = int(num(ws.cell(r, col_rto).value)) if col_rto else 0
        freight = num(ws.cell(r, col_freight).value) if col_freight else 0
        in_transit = int(num(ws.cell(r, col_intransit).value)) if col_intransit else 0
        cancelled = max(0, orders - shipped)

        products[name] = {
            "total_orders": orders, "shipped": shipped, "delivered": delivered,
            "rto": rto, "in_transit": in_transit, "cancelled": cancelled, "lost": 0,
            "revenue": round(revenue, 2), "freight": round(freight, 2),
        }

    return products, round(total_ad, 2)


def parse_amazon_section(ws, start_row, end_row):
    """Parse Amazon section — columns: Products, Total revenue, Total expense, Total P/L,
    Profit %, P/pcs, Total orders, Product expense, Delivered, Return, Amazon expense, RTO%, Only Ad spend"""
    header_row = find_header_row(ws, start_row, end_row)
    if not header_row:
        return {}, 0

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h and isinstance(h, str):
            headers[h.strip()] = c

    col_rev = headers.get('Total revenue', headers.get('Total Revenue', None))
    col_expense = headers.get('Total expense', None)
    col_orders = headers.get('Total orders', None)
    col_prodexp = headers.get('Product expense', None)
    col_delivered = headers.get('Delivered', None)
    col_return = headers.get('Return', None)
    col_amzexp = headers.get('Amazon expense', headers.get('Amazon fees', None))
    col_adspend = headers.get('Only Ad spend', None)

    total_ad = 0
    for r in range(header_row, end_row + 1):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str) and v.strip() == 'Marketing':
            total_ad = num(ws.cell(r, 2).value)
            break

    products = {}
    for r in range(header_row + 1, end_row + 1):
        name = ws.cell(r, 1).value
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if name in SKIP_NAMES or name.startswith('Total'): break
        name = NAME_NORM.get(name, name)

        revenue = num(ws.cell(r, col_rev).value) if col_rev else 0
        orders = int(num(ws.cell(r, col_orders).value)) if col_orders else 0
        if revenue == 0 and orders == 0: continue

        delivered = int(num(ws.cell(r, col_delivered).value)) if col_delivered else 0
        returns = int(num(ws.cell(r, col_return).value)) if col_return else 0
        cancelled = max(0, orders - delivered - returns)
        amz_expense = num(ws.cell(r, col_amzexp).value) if col_amzexp else 0
        ad_spend = num(ws.cell(r, col_adspend).value) if col_adspend else 0

        # For Amazon: shipped ≈ delivered + returns (FBA)
        shipped = delivered + returns

        # Amazon fees = commission + FBA + closing (combined in 'Amazon expense')
        products[name] = {
            "total_orders": orders, "shipped": shipped, "delivered": delivered,
            "rto": returns, "in_transit": 0, "cancelled": cancelled, "lost": 0,
            "revenue": round(revenue, 2), "freight": 0,
            "commission": 0, "fba_fees": 0, "closing_fee": 0,
            "promos": 0, "refund_amt": 0,
            "ad_spend": round(ad_spend, 2),
            "amazon_fees_total": round(amz_expense, 2),
        }

    return products, round(total_ad, 2)


def parse_flipkart_section(ws, start_row, end_row):
    """Parse Flipkart section — similar to Amazon."""
    header_row = find_header_row(ws, start_row, end_row)
    if not header_row:
        return {}, 0

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h and isinstance(h, str):
            headers[h.strip()] = c

    col_rev = headers.get('Total revenue', None)
    col_orders = headers.get('Total orders', None)
    col_prodexp = headers.get('Product expense', None)
    col_delivered = headers.get('Delivered', None)
    col_return = headers.get('Return', None)
    col_fkexp = headers.get('Flipkart expenses', headers.get('Flipkart exp', None))
    col_adspend = headers.get('Only Ad spend', None)

    total_ad = 0
    for r in range(header_row, end_row + 1):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str) and v.strip() == 'Marketing':
            total_ad = num(ws.cell(r, 2).value)
            break

    products = {}
    for r in range(header_row + 1, end_row + 1):
        name = ws.cell(r, 1).value
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if name in SKIP_NAMES or name.startswith('Total'): break
        name = NAME_NORM.get(name, name)

        revenue = num(ws.cell(r, col_rev).value) if col_rev else 0
        orders = int(num(ws.cell(r, col_orders).value)) if col_orders else 0
        if revenue == 0 and orders == 0: continue

        delivered = int(num(ws.cell(r, col_delivered).value)) if col_delivered else 0
        returns = int(num(ws.cell(r, col_return).value)) if col_return else 0
        fk_expense = num(ws.cell(r, col_fkexp).value) if col_fkexp else 0
        ad_spend = num(ws.cell(r, col_adspend).value) if col_adspend else 0
        shipped = delivered + returns

        products[name] = {
            "total_orders": orders, "shipped": shipped, "delivered": delivered,
            "rto": returns, "in_transit": 0, "cancelled": max(0, orders - shipped), "lost": 0,
            "revenue": round(revenue, 2), "freight": 0,
            "commission": 0, "fba_fees": 0, "closing_fee": 0,
            "promos": 0, "refund_amt": 0,
            "ad_spend": round(ad_spend, 2),
            "flipkart_fees_total": round(fk_expense, 2),
        }

    return products, round(total_ad, 2)


def parse_firstcry_section(ws, start_row, end_row):
    """Parse FirstCry section."""
    header_row = find_header_row(ws, start_row, end_row)
    if not header_row:
        return {}, 0

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h and isinstance(h, str):
            headers[h.strip()] = c

    col_rev = headers.get('Total revenue', None)
    col_orders = headers.get('Total orders', None)
    col_delivered = headers.get('Delivered', None)
    col_return = headers.get('Return', None)
    col_fcexp = headers.get('FirstCry exp', headers.get('Firstcry exp', None))

    products = {}
    for r in range(header_row + 1, end_row + 1):
        name = ws.cell(r, 1).value
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if name in SKIP_NAMES or name.startswith('Total'): break
        name = NAME_NORM.get(name, name)

        revenue = num(ws.cell(r, col_rev).value) if col_rev else 0
        orders = int(num(ws.cell(r, col_orders).value)) if col_orders else 0
        if revenue == 0 and orders == 0: continue

        delivered = int(num(ws.cell(r, col_delivered).value)) if col_delivered else 0
        returns = int(num(ws.cell(r, col_return).value)) if col_return else 0
        shipped = delivered + returns

        products[name] = {
            "total_orders": orders, "shipped": shipped, "delivered": delivered,
            "rto": returns, "in_transit": 0, "cancelled": max(0, orders - shipped), "lost": 0,
            "revenue": round(revenue, 2), "freight": 0,
        }

    return products, 0


def parse_blinkit_section(ws, start_row, end_row):
    """Parse Blinkit section — PO model: revenue, product exp, orders, delivered, ads, logistics."""
    header_row = find_header_row(ws, start_row, end_row)
    if not header_row:
        return {}, 0

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h and isinstance(h, str):
            headers[h.strip()] = c

    col_rev = headers.get('Total revenue', None)
    col_orders = headers.get('Total orders', None)
    col_delivered = headers.get('Delivered', None)
    col_prodexp = headers.get('Product expense', headers.get('Product exp', None))
    col_ads = headers.get('Ads', None)
    col_logistics = headers.get('Logistics', None)
    col_bkexp = headers.get('Blinkit exp', None)

    # Get total ad from summary
    total_ad = 0
    for r in range(header_row, end_row + 1):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str) and v.strip() == 'Ad Spent':
            total_ad = num(ws.cell(r, 2).value)
            break

    products = {}
    for r in range(header_row + 1, end_row + 1):
        name = ws.cell(r, 1).value
        if not name or not isinstance(name, str): continue
        name = name.strip()
        if name in SKIP_NAMES or name.startswith('Total'): break
        name = NAME_NORM.get(name, name)

        revenue = num(ws.cell(r, col_rev).value) if col_rev else 0
        orders = int(num(ws.cell(r, col_orders).value)) if col_orders else 0
        if revenue == 0 and orders == 0: continue

        delivered = int(num(ws.cell(r, col_delivered).value)) if col_delivered else orders
        ad_spend = num(ws.cell(r, col_ads).value) if col_ads else 0

        products[name] = {
            "total_orders": orders, "shipped": orders, "delivered": delivered,
            "rto": 0, "in_transit": 0, "cancelled": 0, "lost": 0,
            "revenue": round(revenue, 2), "freight": 0,
            "ad_spend": round(ad_spend, 2),
        }

    return products, round(total_ad, 2)


def parse_instamart_section(ws, start_row, end_row):
    """Parse Instamart section — same structure as Blinkit."""
    return parse_blinkit_section(ws, start_row, end_row)


# Channel parser dispatch
PARSERS = {
    'D2C': parse_d2c_section,
    'Amazon': parse_amazon_section,
    'Flipkart': parse_flipkart_section,
    'FirstCry': parse_firstcry_section,
    'Blinkit': parse_blinkit_section,
    'Instamart': parse_instamart_section,
}


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # Store all data by channel and month
    all_channels = {ch: {} for ch in PARSERS}
    all_adspend = {ch: {} for ch in PARSERS}

    for sheet_name, month_label in SHEET_MAP.items():
        ws = wb[sheet_name]
        sections = find_sections(ws)
        print(f"\n{'='*60}")
        print(f"  {month_label} ({sheet_name}): {[s[0] for s in sections]}")
        print(f"{'='*60}")

        for channel, start_row, end_row in sections:
            parser = PARSERS.get(channel)
            if not parser:
                print(f"  SKIP: Unknown channel '{channel}'")
                continue

            products, ad_total = parser(ws, start_row, end_row)
            if products:
                all_channels[channel][month_label] = products
                all_adspend[channel][month_label] = ad_total

                total_rev = sum(p['revenue'] for p in products.values())
                total_orders = sum(p['total_orders'] for p in products.values())
                total_del = sum(p['delivered'] for p in products.values())
                print(f"  {channel}: {len(products)} products | {total_orders} orders | {total_del} delivered | ₹{total_rev/1e5:.2f}L | Ads: ₹{ad_total/1e5:.2f}L")

                # Check COGS coverage
                for pname in products:
                    if pname not in COGS_MAP:
                        print(f"    ⚠ No COGS mapping for: {pname}")

    # Save JSON files per channel
    print(f"\n{'='*60}")
    print(f"  Saving JSON files...")
    print(f"{'='*60}")

    for channel, month_data in all_channels.items():
        for month_label, products in month_data.items():
            prefix = {
                'D2C': '', 'Amazon': 'amazon_', 'Flipkart': 'flipkart_',
                'FirstCry': 'firstcry_', 'Blinkit': 'blinkit_', 'Instamart': 'instamart_'
            }[channel]
            fname = prefix + month_label.lower().replace(' ', '_') + '_mis_data.json'
            with open(fname, 'w') as f:
                json.dump(products, f, indent=2)
            print(f"  Saved: {fname}")

    # Save ad spend JSON
    with open('fy24_25_adspend.json', 'w') as f:
        json.dump(all_adspend, f, indent=2)
    print(f"  Saved: fy24_25_adspend.json")

    # Print grand summary
    print(f"\n{'='*60}")
    print(f"  FY24-25 ALL CHANNELS SUMMARY")
    print(f"{'='*60}")
    for channel in PARSERS:
        if not all_channels[channel]:
            print(f"\n  {channel}: No data")
            continue
        total_rev = sum(sum(p['revenue'] for p in m.values()) for m in all_channels[channel].values())
        total_orders = sum(sum(p['total_orders'] for p in m.values()) for m in all_channels[channel].values())
        months = sorted(all_channels[channel].keys())
        print(f"\n  {channel}: {len(months)} months | {total_orders:,} orders | ₹{total_rev/1e5:.2f}L revenue")
        print(f"    Months: {', '.join(months)}")


if __name__ == '__main__':
    main()

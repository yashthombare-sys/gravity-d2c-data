#!/usr/bin/env python3
"""
Generate FirstCry MIS data (Apr 2025 – Feb 2026).

Source: "Firstcry Data 1.xlsx" (manually compiled by team)
+ Payment Advice files for Feb 2026 (not in manual file)

Revenue = Base Cost (what FirstCry pays us, NOT MRP).
FirstCry margin = MRP - Base Cost (their cut).
"""

import json, os
import openpyxl

BASE = "/Users/yashthombare/Desktop/Gravity/Shiprocket D2C data"
MANUAL_FILE = os.path.join(BASE, "Firstcry Data 1.xlsx")
FC_DIR = os.path.join(BASE, "Firstcry")

# Product name normalization from manual file
NAME_MAP = {
    "V1": "V1", "V2": "V2", "V3": "V3", "V4": "V4",
    "V5": "V5", "V6": "V6", "V7": "V7 Police Cruiser",
    "V8": "V8", "V9": "V9", "V10": "V10",
    "BB": "Busy Book Blue", "BB Blue": "Busy Book Blue",
    "BB Pink": "Busy Book Pink",
    "V1V4": "V1-V4 Combo", "V1V2": "V1-V2 Combo", "V1V6": "V1-V6 Combo",
    "V2V4": "V2-V4 Combo", "V2V6": "V6-V2 Combo", "V6V2": "V6-V2 Combo",
    "V9V10": "V9-V10 Combo", "V4V6": "V4-V6 Combo", "V2V9": "V2-V9 Combo",
    "V4 Pack of 3": "V4- P of 3", "V4 Pack of 2": "V4- P of 2",
    "V1 Pack of 2": "V1- P of 2",
}

COGS_MAP = {
    "V1": 225, "V2": 275, "V3": 662, "V4": 170, "V5": 225,
    "V6": 275, "V7 Police Cruiser": 600, "V8": 700,
    "V9": 778, "V10": 1009,
    "V1- P of 2": 531, "V4- P of 2": 401, "V4- P of 3": 510,
    "V1-V2 Combo": 524, "V1-V4 Combo": 404, "V1-V6 Combo": 500,
    "V2-V4 Combo": 488, "V6-V2 Combo": 612, "V4-V6 Combo": 445,
    "V9-V10 Combo": 1787, "V2-V9 Combo": 1053,
    "Busy Book Blue": 300, "Busy Book Pink": 300,
}

SHEET_TO_MONTH = {
    "April": "Apr 2025", "May": "May 2025", "June": "Jun 2025",
    "July": "Jul 2025", "Aug": "Aug 2025", "Sep": "Sep 2025",
    "OCt": "Oct 2025", "Nov": "Nov 2025", "Dec": "Dec 2025",
    "Jan26": "Jan 2026",
}

# ProductID → Product for Payment Advice (Feb 2026)
PID_MAP = {
    "16683248": "V1", "16683249": "V2", "16683251": "V4",
    "20271602": "V6", "21060002": "V9", "20752697": "V3",
    "20271598": "V3", "20752700": "V10", "20271599": "V8",
    "21930164": "Busy Book Blue", "21930165": "Busy Book Pink",
    "21060003": "V1-V2 Combo", "19914061": "V1-V2 Combo",
    "19914060": "V1-V4 Combo", "19914062": "V1-V6 Combo",
    "19914063": "V6-V2 Combo", "19914064": "V2-V4 Combo",
    "21060009": "V4-V6 Combo", "21060008": "V9-V10 Combo",
    "21930167": "V2-V9 Combo",
    "21060004": "V1- P of 2", "21060006": "V4- P of 2",
    "21060007": "V4- P of 3",
}


def safe_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("₹", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def read_manual_file():
    """Read team's manual MIS file."""
    wb = openpyxl.load_workbook(MANUAL_FILE, data_only=True)
    monthly = {}

    for sheet_name, month_label in SHEET_TO_MONTH.items():
        ws = wb[sheet_name]
        products = {}

        for r in range(2, ws.max_row + 1):
            raw_name = str(ws.cell(r, 1).value or "").strip()
            if not raw_name:
                continue

            product = NAME_MAP.get(raw_name, raw_name)
            rev = safe_float(ws.cell(r, 2).value)
            delivered = int(safe_float(ws.cell(r, 3).value))

            if rev == 0 and delivered == 0:
                continue

            if product in products:
                products[product]["revenue"] += rev
                products[product]["delivered"] += delivered
                products[product]["total_orders"] += delivered
            else:
                products[product] = {
                    "revenue": round(rev, 2),
                    "total_orders": delivered,
                    "delivered": delivered,
                    "returned": 0,
                    "cancelled": 0,
                    "cogs_unit": COGS_MAP.get(product, 0),
                    "fc_margin": 0,
                    "ad_spend": 0,
                }

        monthly[month_label] = products

    wb.close()
    return monthly


def read_feb_from_payment_advice():
    """Read Feb 2026 data from Payment Advice files."""
    from collections import defaultdict

    products = defaultdict(lambda: {
        "revenue": 0.0, "total_orders": 0, "delivered": 0,
        "returned": 0, "cancelled": 0, "cogs_unit": 0,
        "fc_margin": 0, "ad_spend": 0,
    })

    seen_orders = set()
    files = sorted([f for f in os.listdir(FC_DIR) if f.startswith("ExplortPayment")])

    for fn in files:
        wb = openpyxl.load_workbook(os.path.join(FC_DIR, fn))
        ws = wb.active

        header_row = None
        for r in range(1, 10):
            if str(ws.cell(r, 1).value or "").strip() == "Sr No.":
                header_row = r
                break
        if not header_row:
            wb.close()
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            order_id = str(ws.cell(r, 3).value or "").strip()
            if not order_id:
                continue

            order_date = str(ws.cell(r, 4).value or "")
            # Only Feb 2026
            if "/02/2026" not in order_date and "2026-02" not in order_date:
                continue

            if order_id in seen_orders:
                continue
            seen_orders.add(order_id)

            pid = str(ws.cell(r, 8).value or "").strip()
            product = PID_MAP.get(pid)
            if not product:
                continue

            qty = int(safe_float(ws.cell(r, 10).value) or 1)
            base_cost = safe_float(ws.cell(r, 12).value)

            p = products[product]
            p["revenue"] += base_cost * qty
            p["total_orders"] += qty
            p["delivered"] += qty
            p["cogs_unit"] = COGS_MAP.get(product, 0)

        wb.close()

    return dict(products)


def save_json_files(monthly):
    """Save per-month JSON files."""
    output_map = {
        "Apr 2025": "firstcry_apr_2025_mis_data.json",
        "May 2025": "firstcry_may_2025_mis_data.json",
        "Jun 2025": "firstcry_jun_2025_mis_data.json",
        "Jul 2025": "firstcry_jul_2025_mis_data.json",
        "Aug 2025": "firstcry_aug_2025_mis_data.json",
        "Sep 2025": "firstcry_sep_2025_mis_data.json",
        "Oct 2025": "firstcry_oct_2025_mis_data.json",
        "Nov 2025": "firstcry_nov_2025_mis_data.json",
        "Dec 2025": "firstcry_dec_2025_mis_data.json",
        "Jan 2026": "firstcry_jan_2026_mis_data.json",
        "Feb 2026": "firstcry_feb_2026_mis_data.json",
    }

    for month_label, filename in output_map.items():
        products = monthly.get(month_label, {})
        output = {}
        for product, p in sorted(products.items()):
            if p["total_orders"] == 0 and p["revenue"] == 0:
                continue
            output[product] = {
                "revenue": round(p["revenue"], 2),
                "total_orders": p["total_orders"],
                "delivered": p["delivered"],
                "returned": p["returned"],
                "cancelled": p["cancelled"],
                "cogs_unit": p["cogs_unit"],
                "fc_margin": round(p.get("fc_margin", 0), 2),
                "ad_spend": 0,
            }

        filepath = os.path.join(BASE, filename)
        with open(filepath, "w") as f:
            json.dump(output, f, indent=2)

        total_orders = sum(p["total_orders"] for p in output.values())
        total_delivered = sum(p["delivered"] for p in output.values())
        total_rev = sum(p["revenue"] for p in output.values())
        print(f"  {month_label}: {len(output)} products, "
              f"{total_orders} orders, {total_delivered} delivered, "
              f"₹{total_rev:,.0f} revenue → {filename}")


def main():
    print("\n🔄 Generating FirstCry MIS (Apr 2025 – Feb 2026)\n")

    print("📖 Reading manual MIS file (Firstcry Data 1.xlsx)...")
    monthly = read_manual_file()

    print("\n📦 Reading Feb 2026 from Payment Advice files...")
    feb_data = read_feb_from_payment_advice()
    monthly["Feb 2026"] = feb_data

    print("\n💾 Saving JSON files...")
    save_json_files(monthly)

    # Summary
    print("\n" + "=" * 65)
    print("📊 FIRSTCRY MIS SUMMARY")
    print("=" * 65)
    grand_orders = 0
    grand_revenue = 0
    for month in ["Apr 2025", "May 2025", "Jun 2025", "Jul 2025", "Aug 2025",
                   "Sep 2025", "Oct 2025", "Nov 2025", "Dec 2025", "Jan 2026", "Feb 2026"]:
        products = monthly.get(month, {})
        orders = sum(p["total_orders"] for p in products.values())
        rev = sum(p["revenue"] for p in products.values())
        grand_orders += orders
        grand_revenue += rev
        print(f"  {month:12s}: {orders:4d} delivered | ₹{rev:>10,.0f}")

    print(f"\n  TOTAL:        {grand_orders:4d} delivered | ₹{grand_revenue:>10,.0f}")
    print(f"\n✅ Done!")


if __name__ == "__main__":
    main()

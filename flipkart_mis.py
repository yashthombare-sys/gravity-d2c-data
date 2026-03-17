#!/usr/bin/env python3
"""
Generate Flipkart MIS data (Apr 2025 – Feb 2026).

Reads:
  - Settlement xlsx files (Q1-Q3 quarterly + Jan/Feb monthly)
  - Fulfilment report (order statuses)
  - Ad spend CSVs (daily reports from Flipkart Ads)

Outputs:
  - flipkart_{month}_mis_data.json per month
"""

import json, os, re
from collections import defaultdict
from datetime import datetime
import openpyxl

BASE = "/Users/yashthombare/Desktop/Gravity/Shiprocket D2C data"
FK_DIR = os.path.join(BASE, "Flipkart Data")

# ── SKU → Product mapping ──
SKU_MAP = {
    "v4_new": "V4",
    "v1_new": "V1",
    "V2": "V2",
    "v6": "V6",
    "Toy- V9": "V9",
    "CLAPSTORE-V9": "V9",
    "Relist - V9": "V9",
    "COMBO_V1V4": "V1-V4 Combo",
    "COMBO_V1V2": "V1-V2 Combo",
    "COMBO_V4V2": "V2-V4 Combo",
    "COMBO_V6V2": "V6-V2 Combo",
    "COMBO_V1V6": "V6-V1 Combo",
    "COMBO_V2V9": "V9-V2 Combo",
    "wooden Calculator": "Drawing Board",
    "v7": "V7 Police Cruiser",
    "RELIST - V3": "V3",
    "v3": "V3",
    "V10": "V10",
    "busybook": "Busy Book Blue",
    "v5": "V5",
    "V4 _PACK OF 2": "V4- P of 2",
    "TOY-V8": "V8",
    "Color Matching Game": "Color Matching Game",
    # Shopsy variants → same product
    "SH-TOY-V6": "V6",
    "SH-TOY-V5": "V5",
    "SH-TOY-COMBO_V4V2": "V2-V4 Combo",
    "SH-TOY-COMBO_V6V2": "V6-V2 Combo",
    # V1 + Calculator combo listing
    "V1 & Caculator": "V1-Calculator Combo",
    # Review SKUs — likely zero-price review bait, will be filtered by revenue
    "REVIEWS FOR V7 - CALCULATOR": "Drawing Board",
    "REVIEWS FOR V4_NEW - CALCULATOR": "Drawing Board",
    "REVIEW FOR V2 - BUSYBOOK": "Busy Book Blue",
    "Reviews - Busybook": "Busy Book Blue",
    # Duplicate listing
    "DUP TOY-V4": "V4",
    "seller_easy_ship": None,  # This is fulfilment type, not SKU
}

# COGS per unit (same as D2C/Amazon)
COGS_MAP = {
    "V1": 225, "V2": 275, "V3": 662, "V4": 170, "V5": 225,
    "V6": 275, "V7 Police Cruiser": 600, "V8": 700,
    "V9": 778, "V10": 1009,
    "V1- P of 2": 531, "V1- P of 3": 675, "V2- P of 2": 649,
    "V4- P of 2": 401, "V6- P of 2": 649, "V9- P of 2": 1664,
    "V6-V1 Combo": 608, "V6-V2 Combo": 612,
    "V1-V2 Combo": 524, "V1-V4 Combo": 404, "V2-V4 Combo": 488,
    "V9-V2 Combo": 488, "V9-V3 Combo": 1440, "V9-V10 Combo": 1787,
    "V6-V1 Combo": 608,
    "Busy Book Blue": 300, "Busy Book Pink": 300, "Human Book": 300,
    "Ganesha": 290, "Krishna": 290, "Hanuman": 290,
    "Car": 540, "Tank": 862, "JCB": 540, "Drawing Board": 250,
    "Color Matching Game": 200,
    "V1-Calculator Combo": 475,  # V1(225) + Drawing Board(250)
}

MONTH_MAP = {
    "2025-04": "Apr 2025", "2025-05": "May 2025", "2025-06": "Jun 2025",
    "2025-07": "Jul 2025", "2025-08": "Aug 2025", "2025-09": "Sep 2025",
    "2025-10": "Oct 2025", "2025-11": "Nov 2025", "2025-12": "Dec 2025",
    "2026-01": "Jan 2026", "2026-02": "Feb 2026",
}

# ── Settlement files ──
SETTLEMENT_FILES = [
    "ef0aa7f67b4e477b_Q1_FY_25_26.xlsx",
    "ef0aa7f67b4e477b_Q2_FY_25_26.xlsx",
    "ef0aa7f67b4e477b_Q3_FY_25_26.xlsx",
    "727eb77b-0c5e-4b17-8673-5d426428531b_1773317760000.xlsx",
    "ad236e91-a179-48dc-bb0c-aa964d50bd06_1773317739000.xlsx",
]

FULFILMENT_FILE = "f773a298-3e3d-4aa0-8d40-d1a31e068c42_1773316944000.xlsx"


def find_col(ws, header_row, name):
    """Find column index by header name (partial match)."""
    for j in range(1, ws.max_column + 1):
        h = str(ws.cell(header_row, j).value or "")
        if name.lower() in h.lower():
            return j
    return None


def safe_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("₹", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def map_sku(raw_sku):
    """Map raw Flipkart SKU to product name."""
    sku = str(raw_sku).strip().strip('"')
    if sku in SKU_MAP:
        return SKU_MAP[sku]
    # Try case-insensitive
    for k, v in SKU_MAP.items():
        if k.lower() == sku.lower():
            return v
    return sku  # return as-is if no mapping


def read_settlement_files():
    """Read all settlement xlsx files. Aggregate by (order_item_id, month) → product data."""
    # Per order_item: aggregate sale, fees, refund
    order_items = defaultdict(lambda: {
        "sale_amount": 0.0, "offer_amount": 0.0,
        "commission": 0.0, "fixed_fee": 0.0, "collection_fee": 0.0,
        "shipping_fee": 0.0, "reverse_shipping_fee": 0.0,
        "marketplace_fee": 0.0, "refund": 0.0,
        "sku": "", "product": "", "quantity": 0, "order_date": "",
        "month": "",
    })

    for fn in SETTLEMENT_FILES:
        path = os.path.join(FK_DIR, fn)
        if not os.path.exists(path):
            print(f"  ⚠️  Missing: {fn}")
            continue

        print(f"  Reading {fn}...")
        wb = openpyxl.load_workbook(path)
        ws = wb["Orders"]

        # Find columns by header (row 2)
        cols = {}
        for j in range(1, ws.max_column + 1):
            h = str(ws.cell(2, j).value or "").strip()
            hl = h.lower()
            if "order item id" in hl:
                cols["item_id"] = j
            elif "order id" in hl and "item" not in hl:
                cols["order_id"] = j
            elif h == "Sale Amount (Rs.)":
                cols["sale"] = j
            elif "total offer amount" in hl and j < 20:
                cols["offer"] = j
            elif "marketplace fee" in hl:
                cols["mp_fee"] = j
            elif h == "Refund (Rs.)":
                cols["refund"] = j
            elif h == "Commission (Rs.)":
                cols["commission"] = j
            elif "fixed fee" in hl:
                cols["fixed_fee"] = j
            elif "collection fee" in hl:
                cols["collection_fee"] = j
            elif h == "Pick And Pack Fee (Rs.)":
                cols["pick_pack"] = j
            elif h == "Shipping Fee (Rs.)":
                cols["shipping"] = j
            elif h == "Reverse Shipping Fee (Rs.)":
                cols["reverse_shipping"] = j
            elif "seller sku" in hl:
                cols["sku"] = j
            elif h == "Quantity" or hl == "quantity":
                cols["quantity"] = j
            elif "order date" in hl:
                cols["order_date"] = j

        print(f"    Columns found: {list(cols.keys())}")

        for r in range(4, ws.max_row + 1):
            item_id = str(ws.cell(r, cols["item_id"]).value or "").strip()
            if not item_id:
                continue

            raw_sku = str(ws.cell(r, cols["sku"]).value or "").strip().strip('"')
            product = map_sku(raw_sku)
            if product is None:
                continue  # Skip non-product rows

            order_date = str(ws.cell(r, cols["order_date"]).value or "")[:10]
            month_key = order_date[:7]  # YYYY-MM

            key = (item_id, month_key)
            d = order_items[key]
            d["sale_amount"] += safe_float(ws.cell(r, cols["sale"]).value)
            d["offer_amount"] += safe_float(ws.cell(r, cols.get("offer", 0)).value) if cols.get("offer") else 0
            d["commission"] += abs(safe_float(ws.cell(r, cols["commission"]).value))
            d["fixed_fee"] += abs(safe_float(ws.cell(r, cols["fixed_fee"]).value))
            d["collection_fee"] += abs(safe_float(ws.cell(r, cols.get("collection_fee", 0)).value)) if cols.get("collection_fee") else 0
            d["shipping_fee"] += abs(safe_float(ws.cell(r, cols["shipping"]).value))
            d["reverse_shipping_fee"] += abs(safe_float(ws.cell(r, cols["reverse_shipping"]).value))
            d["marketplace_fee"] += abs(safe_float(ws.cell(r, cols["mp_fee"]).value))
            d["refund"] += abs(safe_float(ws.cell(r, cols["refund"]).value))
            d["sku"] = raw_sku
            d["product"] = product
            d["quantity"] = max(d["quantity"], int(safe_float(ws.cell(r, cols["quantity"]).value)))
            d["order_date"] = order_date
            d["month"] = month_key

        wb.close()

    print(f"  Total order items: {len(order_items)}")
    return order_items


def read_fulfilment_report():
    """Read fulfilment report for order statuses."""
    path = os.path.join(FK_DIR, FULFILMENT_FILE)
    if not os.path.exists(path):
        print("  ⚠️  Fulfilment report not found")
        return {}

    print(f"  Reading fulfilment report...")
    wb = openpyxl.load_workbook(path)
    ws = wb["Orders"]

    # Headers in row 1
    cols = {}
    for j in range(1, ws.max_column + 1):
        h = str(ws.cell(1, j).value or "").strip().lower()
        if h == "order_item_id":
            cols["item_id"] = j
        elif h == "order_id":
            cols["order_id"] = j
        elif h == "order_item_status":
            cols["status"] = j
        elif h == "sku":
            cols["sku"] = j
        elif h == "quantity":
            cols["quantity"] = j
        elif h == "order_date":
            cols["order_date"] = j
        elif h == "order_delivery_date":
            cols["delivery_date"] = j

    statuses = {}  # order_item_id → status
    for r in range(2, ws.max_row + 1):
        item_id_raw = str(ws.cell(r, cols["item_id"]).value or "").strip().strip('"')
        # Remove "OI:" prefix
        item_id = item_id_raw.replace("OI:", "")
        status = str(ws.cell(r, cols["status"]).value or "").strip()
        order_date = str(ws.cell(r, cols["order_date"]).value or "")[:10]
        month_key = order_date[:7]

        raw_sku = str(ws.cell(r, cols["sku"]).value or "").strip().strip('"').replace("SKU:", "")
        product = map_sku(raw_sku)

        statuses[(item_id, month_key)] = {
            "status": status,
            "product": product,
            "order_date": order_date,
        }

    wb.close()
    print(f"  Fulfilment entries: {len(statuses)}")
    return statuses


def build_monthly_data(order_items, fulfilment_statuses):
    """Build per-month, per-product aggregated data."""
    monthly = {}

    for (item_id, month_key), d in order_items.items():
        if month_key not in MONTH_MAP:
            continue

        month_label = MONTH_MAP[month_key]
        if month_label not in monthly:
            monthly[month_label] = {}

        product = d["product"]
        if not product:
            continue

        if product not in monthly[month_label]:
            monthly[month_label][product] = {
                "total_orders": 0,
                "delivered": 0,
                "returned": 0,
                "cancelled": 0,
                "revenue": 0.0,
                "refund_amt": 0.0,
                "commission": 0.0,
                "fixed_fee": 0.0,
                "collection_fee": 0.0,
                "shipping_fee": 0.0,
                "reverse_shipping_fee": 0.0,
                "total_flipkart_fees": 0.0,
                "cogs_unit": COGS_MAP.get(product, 0),
                "ad_spend": 0.0,
            }

        p = monthly[month_label][product]
        qty = max(d["quantity"], 1)

        # Check fulfilment status
        status_info = fulfilment_statuses.get((item_id, month_key))
        if status_info:
            status = status_info["status"]
        else:
            # Estimate from settlement: refund > 0 → returned
            if d["refund"] > 0:
                status = "RETURNED"
            elif d["sale_amount"] == 0 and d["refund"] == 0 and d["marketplace_fee"] == 0:
                status = "CANCELLED"
            else:
                status = "DELIVERED"

        p["total_orders"] += qty

        if status == "DELIVERED":
            p["delivered"] += qty
            p["revenue"] += d["sale_amount"]
        elif status in ("RETURNED", "RETURN_REQUESTED"):
            p["returned"] += qty
            p["refund_amt"] += d["refund"]
            # Returns still have fees charged
        elif status in ("CANCELLED", "REJECTED"):
            p["cancelled"] += qty
        else:
            # READY_TO_SHIP or unknown → count as order, revenue if present
            p["delivered"] += qty
            p["revenue"] += d["sale_amount"]

        p["commission"] += d["commission"]
        p["fixed_fee"] += d["fixed_fee"]
        p["collection_fee"] += d["collection_fee"]
        p["shipping_fee"] += d["shipping_fee"]
        p["reverse_shipping_fee"] += d["reverse_shipping_fee"]
        p["total_flipkart_fees"] += d["marketplace_fee"]

    return monthly


def read_ad_spend():
    """Read ad spend from CSVs (Consolidated Daily Reports)."""
    ad_totals = defaultdict(float)

    for fn in os.listdir(FK_DIR):
        if not fn.endswith(".csv"):
            continue
        path = os.path.join(FK_DIR, fn)
        with open(path) as f:
            lines = f.readlines()

        if len(lines) <= 3:
            # Empty report — extract month from header
            continue

        # Parse header for date range
        # Line 0: "Start Time, YYYY-MM-DD HH:MM:SS"
        start_date = lines[0].split(",")[1].strip().split(" ")[0]
        month_key = start_date[:7]
        month_label = MONTH_MAP.get(month_key)
        if not month_label:
            continue

        # Data rows start at line 3 (line 2 is column headers)
        for line in lines[3:]:
            parts = line.strip().split(",")
            if len(parts) >= 4:
                try:
                    spend = float(parts[3])
                    ad_totals[month_label] += spend
                except (ValueError, IndexError):
                    pass

    return dict(ad_totals)


def finalize_and_save(monthly, ad_spend):
    """Round numbers, calculate net fields, save JSONs."""
    output_files = {
        "Apr 2025": "flipkart_apr_2025_mis_data.json",
        "May 2025": "flipkart_may_2025_mis_data.json",
        "Jun 2025": "flipkart_jun_2025_mis_data.json",
        "Jul 2025": "flipkart_jul_2025_mis_data.json",
        "Aug 2025": "flipkart_aug_2025_mis_data.json",
        "Sep 2025": "flipkart_sep_2025_mis_data.json",
        "Oct 2025": "flipkart_oct_2025_mis_data.json",
        "Nov 2025": "flipkart_nov_2025_mis_data.json",
        "Dec 2025": "flipkart_dec_2025_mis_data.json",
        "Jan 2026": "flipkart_jan_2026_mis_data.json",
        "Feb 2026": "flipkart_feb_2026_mis_data.json",
    }

    for month_label, filename in output_files.items():
        products = monthly.get(month_label, {})
        month_ad = ad_spend.get(month_label, 0)

        # Allocate ad spend proportionally by revenue
        total_rev = sum(p["revenue"] for p in products.values())

        output = {}
        for product, p in sorted(products.items()):
            if p["total_orders"] == 0:
                continue

            # Net revenue = sale amount (already only for delivered)
            net_revenue = round(p["revenue"], 2)

            # Allocate ad spend by revenue share
            if total_rev > 0 and month_ad > 0:
                p["ad_spend"] = round(month_ad * (p["revenue"] / total_rev), 2)

            output[product] = {
                "revenue": net_revenue,
                "total_orders": p["total_orders"],
                "delivered": p["delivered"],
                "returned": p["returned"],
                "cancelled": p["cancelled"],
                "cogs_unit": p["cogs_unit"],
                "commission": round(p["commission"], 2),
                "fixed_fee": round(p["fixed_fee"], 2),
                "collection_fee": round(p["collection_fee"], 2),
                "shipping_fee": round(p["shipping_fee"], 2),
                "reverse_shipping_fee": round(p["reverse_shipping_fee"], 2),
                "total_flipkart_fees": round(p["total_flipkart_fees"], 2),
                "refund_amt": round(p["refund_amt"], 2),
                "ad_spend": round(p.get("ad_spend", 0), 2),
            }

        filepath = os.path.join(BASE, filename)
        with open(filepath, "w") as f:
            json.dump(output, f, indent=2)

        total_orders = sum(p["total_orders"] for p in output.values())
        total_delivered = sum(p["delivered"] for p in output.values())
        total_rev = sum(p["revenue"] for p in output.values())
        print(f"  {month_label}: {len(output)} products, "
              f"{total_orders} orders, {total_delivered} delivered, "
              f"₹{total_rev:,.0f} revenue → {filename}")

    return output_files


def main():
    print("\n🔄 Generating Flipkart MIS (Apr 2025 – Feb 2026)\n")

    print("📦 Reading settlement files...")
    order_items = read_settlement_files()

    print("\n📋 Reading fulfilment report...")
    fulfilment = read_fulfilment_report()

    print("\n📊 Reading ad spend...")
    ad_spend = read_ad_spend()
    if ad_spend:
        for m, v in sorted(ad_spend.items()):
            print(f"  {m}: ₹{v:,.0f}")
    else:
        print("  No ad spend data found (CSVs are empty)")

    print("\n🔧 Building monthly data...")
    monthly = build_monthly_data(order_items, fulfilment)

    print("\n💾 Saving JSON files...")
    finalize_and_save(monthly, ad_spend)

    # Print summary
    print("\n" + "=" * 60)
    print("📊 FLIPKART MIS SUMMARY")
    print("=" * 60)
    grand_orders = 0
    grand_delivered = 0
    grand_revenue = 0
    grand_returned = 0
    for month in MONTH_MAP.values():
        if month in monthly:
            products = monthly[month]
            orders = sum(p["total_orders"] for p in products.values())
            delivered = sum(p["delivered"] for p in products.values())
            returned = sum(p["returned"] for p in products.values())
            rev = sum(p["revenue"] for p in products.values())
            fees = sum(p["total_flipkart_fees"] for p in products.values())
            grand_orders += orders
            grand_delivered += delivered
            grand_returned += returned
            grand_revenue += rev
            print(f"  {month:12s}: {orders:4d} orders | {delivered:4d} delivered | "
                  f"{returned:3d} returned | ₹{rev:>10,.0f} rev | ₹{fees:>8,.0f} fees")

    print(f"\n  TOTAL:        {grand_orders:4d} orders | {grand_delivered:4d} delivered | "
          f"{grand_returned:3d} returned | ₹{grand_revenue:>10,.0f} revenue")
    print(f"\n✅ Done! JSON files saved to {BASE}")


if __name__ == "__main__":
    main()

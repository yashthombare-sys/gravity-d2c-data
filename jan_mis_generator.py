#!/usr/bin/env python3
"""
January 2026 MIS Generator for Shiprocket D2C Data
Fetches orders via Shiprocket API, processes per MIS flow, generates Excel.
"""

import requests
import json
import re
import os
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

# ── Config ──────────────────────────────────────────────────────────────────
BASE_DIR = "/Users/yashthombare/Desktop/Gravity/Shiprocket D2C data"
API_BASE = "https://apiv2.shiprocket.in/v1/external"
JAN_START = "2026-01-01"
JAN_END = "2026-01-31"

# Load token from .env
with open(os.path.join(BASE_DIR, ".env")) as f:
    for line in f:
        if line.startswith("SHIPROCKET_API_TOKEN="):
            TOKEN = line.strip().split("=", 1)[1]

HEADERS = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {TOKEN}"
}

# ── Status Classification ───────────────────────────────────────────────────
DELIVERED_STATUSES = {"DELIVERED"}
RTO_STATUSES = {"RTO DELIVERED", "RTO IN TRANSIT", "RTO INITIATED", "RTO OFD",
                "REACHED BACK AT SELLER CITY", "REACHED BACK AT_SELLER_CITY"}
CANCELLED_STATUSES = {"CANCELED", "CANCELLATION REQUESTED"}
SKIP_STATUSES = {"SELF FULFILED", "QC FAILED", "RETURN DELIVERED",
                 "RETURN IN TRANSIT", "RETURN PENDING", "RETURN CANCELLED"}
# Everything else = In-Transit

SPARE_PARTS_KEYWORDS = [
    "RC TANK MOTOR", "RC CAR PCB", "Charging cable", "Documents",
    "charging cable", "spare", "SPARE", "pcb", "motor", "document"
]

# ── Product Mapping ─────────────────────────────────────────────────────────
# Order matters: check combos first, then longer names, then shorter
PRODUCT_PATTERNS = [
    # Combos first
    (r"V9.*V10|V10.*V9|V9\+V10|V9-V10|V9 \+ V10|V9 V10 Combo", "V9-V10 Combo"),
    (r"V9.*V3|V3.*V9|V9\+V3|V9-V3|V9 \+ V3|V9 V3 Combo", "V9-V3 Combo"),
    (r"V9.*V2|V2.*V9|V9\+V2|V9-V2|V9 \+ V2|V9 V2 Combo", "V9-V2 Combo"),
    (r"V6.*V1|V1.*V6|V6\+V1|V6-V1|V6 \+ V1|V6 V1 Combo", "V6-V1 Combo"),
    (r"V6.*V2|V2.*V6|V6\+V2|V6-V2|V6 \+ V2|V6 V2 Combo", "V6-V2 Combo"),
    (r"V1.*V4|V4.*V1|V1\+V4|V1-V4|V1 \+ V4|V1 V4 Combo", "V1-V4 Combo"),
    (r"V1.*V2|V2.*V1|V1\+V2|V1-V2|V1 \+ V2|V1 V2 Combo", "V1-V2 Combo"),
    (r"V2.*V4|V4.*V2|V2\+V4|V2-V4|V2 \+ V4|V2 V4 Combo", "V2-V4 Combo"),

    # Pack of 2/3 variants
    (r"V9.*(?:Pack of 2|P of 2|pack of 2|2\s*pack)", "V9 P of 2"),
    (r"V6.*(?:Pack of 2|P of 2|pack of 2|2\s*pack)", "V6- P of 2"),
    (r"V4.*(?:Pack of 3|P of 3|pack of 3|3\s*pack)", "V4- P of 3"),
    (r"V4.*(?:Pack of 2|P of 2|pack of 2|2\s*pack)", "V4- P of 2"),
    (r"V2.*(?:Pack of 2|P of 2|pack of 2|2\s*pack)", "V2- P of 2"),
    (r"V1.*(?:Pack of 2|P of 2|pack of 2|2\s*pack)", "V1- P of 2"),

    # Individual products (check V10 before V1)
    (r"V10(?!\d)", "V10"),
    (r"V1(?!\d)", "V1"),
    (r"V2(?!\d)", "V2"),
    (r"V3(?!\d)", "V3"),
    (r"V4(?!\d)", "V4"),
    (r"V6(?!\d)", "V6"),
    (r"V9(?!\d)", "V9"),

    # Busy Books
    (r"(?i)busy\s*book.*(?:blue|boy)", "Busy Book Blue"),
    (r"(?i)busy\s*book.*(?:pink|girl)", "Busy Book Pink"),
    (r"(?i)human\s*(?:body\s*)?(?:busy\s*)?book", "Human Book"),

    # ClapCuddles
    (r"(?i)(?:clap\s*cuddle|clapcuddle).*(?:ganesh|ganesha)", "Ganesha"),
    (r"(?i)(?:clap\s*cuddle|clapcuddle).*(?:krishna)", "Krishna"),
    (r"(?i)(?:clap\s*cuddle|clapcuddle).*(?:hanuman)", "Hanuman"),
    (r"(?i)ganesh", "Ganesha"),
    (r"(?i)krishna", "Krishna"),
    (r"(?i)hanuman", "Hanuman"),

    # RC vehicles
    (r"(?i)(?:sooper\s*brains?\s*)?(?:rc\s*)?(?:army\s*)?tank", "Tank"),
    (r"(?i)(?:sooper\s*brains?\s*)?(?:rc\s*)?(?:racer|car)", "Car"),
    (r"(?i)(?:sooper\s*brains?\s*)?jcb", "JCB"),
]


def classify_product(product_name):
    """Map product name to MIS category."""
    if not product_name:
        return None
    for pattern, category in PRODUCT_PATTERNS:
        if re.search(pattern, product_name):
            return category
    return None


def classify_status(status):
    """Classify order status."""
    if not status:
        return "unknown"
    s = status.upper().strip()
    if s in DELIVERED_STATUSES:
        return "delivered"
    if s in RTO_STATUSES:
        return "rto"
    if s in CANCELLED_STATUSES:
        return "cancelled"
    if s in SKIP_STATUSES or s.startswith("RETURN"):
        return "skip"
    return "in_transit"


def is_spare_part(product_name):
    """Check if product is a spare part."""
    if not product_name:
        return True
    name_lower = product_name.lower()
    for kw in SPARE_PARTS_KEYWORDS:
        if kw.lower() in name_lower:
            return True
    return False


# ── Step 1: Fetch orders from API ───────────────────────────────────────────
def fetch_january_orders():
    """Fetch all orders in January 2026 date range."""
    all_orders = []
    page = 1
    per_page = 200

    print("Fetching January 2026 orders from Shiprocket API...")

    while True:
        url = f"{API_BASE}/orders?per_page={per_page}&page={page}&from={JAN_START}&to={JAN_END}"
        print(f"  Page {page}...", end=" ", flush=True)

        resp = requests.get(url, headers=HEADERS)
        if resp.status_code == 401:
            print("\nERROR: Token expired. Re-authenticate and update .env")
            return None

        if resp.status_code != 200:
            print(f"\nERROR: API returned {resp.status_code}: {resp.text[:200]}")
            return None

        data = resp.json()

        # Handle response structure
        if isinstance(data, dict) and "data" in data:
            orders = data["data"]
            if isinstance(orders, dict):
                orders = orders.get("data", orders.get("orders", []))
            if not isinstance(orders, list):
                orders = []
        elif isinstance(data, list):
            orders = data
        else:
            orders = []

        if not orders:
            print(f"0 orders (done)")
            break

        all_orders.extend(orders)
        print(f"{len(orders)} orders")

        # Check if there are more pages
        meta = data.get("meta", {}) if isinstance(data, dict) else {}
        pagination = meta.get("pagination", {})
        total_pages = pagination.get("total_pages", 0)

        if isinstance(data, dict) and "data" in data and isinstance(data["data"], dict):
            last_page = data["data"].get("last_page", 0)
            if last_page and page >= last_page:
                break

        if total_pages and page >= total_pages:
            break

        if len(orders) < per_page:
            break

        page += 1

    print(f"\nTotal orders fetched: {len(all_orders)}")
    return all_orders


# ── Step 2: Load freight data ───────────────────────────────────────────────
def load_freight_data():
    """Load freight from Freight Total Amount.xlsx, deduplicated by Order ID."""
    freight_file = os.path.join(BASE_DIR, "Freight Total Amount.xlsx")
    wb = openpyxl.load_workbook(freight_file, read_only=True, data_only=True)
    ws = wb.active

    # Find header row
    headers = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col_idx

    # Try different possible column names
    order_col = None
    freight_col = None
    for name in ["order id", "orderid", "order_id", "sr order id"]:
        if name in headers:
            order_col = headers[name]
            break
    for name in ["freight total amount", "freight", "total freight", "freight total"]:
        if name in headers:
            freight_col = headers[name]
            break

    if not order_col or not freight_col:
        print(f"WARNING: Could not find columns in freight file. Headers: {headers}")
        # Try first two columns as fallback
        order_col = 1
        freight_col = 2

    freight_map = {}
    for row in ws.iter_rows(min_row=2):
        order_id = row[order_col - 1].value
        freight_val = row[freight_col - 1].value
        if order_id and freight_val:
            order_id = str(order_id).strip()
            if order_id not in freight_map:  # Deduplicate: keep first
                try:
                    freight_map[order_id] = float(freight_val)
                except (ValueError, TypeError):
                    pass

    wb.close()
    print(f"Loaded freight for {len(freight_map)} unique orders")
    return freight_map


# ── Step 3: Process orders ──────────────────────────────────────────────────
def process_orders(orders, freight_map):
    """Process orders into MIS data structure."""
    # Track per-product data
    product_data = defaultdict(lambda: {
        "total_orders": 0,
        "shipped": 0,
        "delivered": 0,
        "rto": 0,
        "in_transit": 0,
        "cancelled": 0,
        "revenue": 0.0,
        "freight": 0.0,
    })

    seen = set()  # Deduplicate by (order_id, product_name)
    unmapped = defaultdict(int)
    skipped_statuses = 0
    spare_parts_skipped = 0

    for order in orders:
        order_id = str(order.get("id", order.get("order_id", "")))
        channel_order_id = str(order.get("channel_order_id", order_id))
        is_reverse = order.get("is_reverse", False)

        # Skip reverse orders
        if is_reverse or str(is_reverse).lower() == "yes" or str(is_reverse) == "1":
            continue

        status_raw = order.get("status", order.get("status_code", ""))
        status = classify_status(status_raw)

        if status == "skip":
            skipped_statuses += 1
            continue

        # Get products from order
        products = order.get("products", order.get("order_items", []))
        if not products:
            # Single product order structure
            product_name = order.get("product_name", "")
            if product_name:
                products = [{
                    "name": product_name,
                    "selling_price": order.get("product_price", order.get("selling_price", 0)),
                    "discount": order.get("discount", 0),
                    "quantity": order.get("product_quantity", order.get("quantity", 1)),
                }]

        # Calculate total order value for freight allocation
        order_line_values = []
        order_products_info = []

        for prod in products:
            pname = prod.get("name", prod.get("product_name", ""))
            # API: actual price is in 'price', selling_price is often 0
            price = float(prod.get("price", prod.get("selling_price", prod.get("product_price", 0))) or 0)
            discount = float(prod.get("discount", 0) or 0)
            qty = int(prod.get("quantity", prod.get("product_quantity", 1)) or 1)

            # Skip spare parts
            if is_spare_part(pname):
                spare_parts_skipped += 1
                continue

            # Skip fake quantity
            if qty > 10:
                continue

            category = classify_product(pname)
            if not category:
                unmapped[pname] += 1
                continue

            dedup_key = (order_id, category)
            if dedup_key in seen:
                continue
            seen.add(dedup_key)

            line_value = (price - discount) * qty
            order_line_values.append(line_value)
            order_products_info.append((category, price, discount, qty, line_value))

        # Allocate freight proportionally
        total_order_value = sum(order_line_values) if order_line_values else 0
        # Freight file uses channel_order_id (e.g. "CS76517"), try that first
        order_freight = freight_map.get(channel_order_id, freight_map.get(order_id, 0))

        for i, (category, price, discount, qty, line_value) in enumerate(order_products_info):
            pd = product_data[category]
            pd["total_orders"] += 1

            if status == "delivered":
                pd["delivered"] += 1
                pd["shipped"] += 1
                pd["revenue"] += line_value
            elif status == "rto":
                pd["rto"] += 1
                pd["shipped"] += 1
            elif status == "in_transit":
                pd["in_transit"] += 1
                pd["shipped"] += 1
            elif status == "cancelled":
                pd["cancelled"] += 1

            # Proportional freight
            if total_order_value > 0 and order_freight > 0:
                freight_share = (line_value / total_order_value) * order_freight
                pd["freight"] += freight_share

    if unmapped:
        print(f"\nUnmapped products (top 20):")
        for name, count in sorted(unmapped.items(), key=lambda x: -x[1])[:20]:
            print(f"  {name}: {count}")

    print(f"Spare parts skipped: {spare_parts_skipped}")
    print(f"Skipped statuses: {skipped_statuses}")

    return dict(product_data)


# ── Step 4: Generate Excel MIS ──────────────────────────────────────────────
def generate_mis_excel(product_data):
    """Generate the MIS Excel sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "January 2026 MIS"

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    currency_fmt = '#,##0'
    pct_fmt = '0.0%'
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # MIS columns
    col_headers = [
        "Products",                  # A
        "Total Delivered Revenue",   # B
        "Total Expense",             # C (=I+K)
        "Total P/L",                 # D (=B-C)
        "Profit %",                  # E (=D/B)
        "P/pcs",                     # F (=D/J)
        "Total Orders",              # G
        "Shipped",                   # H
        "Total COGS",               # I (=R*H)
        "Delivered",                 # J
        "Shipping Charges",          # K
        "RTO",                       # L
        "In-Transit",               # M
        "RTO%",                      # N (=L/H)
        "Shipped%",                  # O (=H/G)
        "Delivered%",               # P (=J/G)
        "Cancellation%",            # Q (=(G-H)/G)
        "COGS/Unit",                # R
    ]

    # Write headers
    for col, header in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    # Product order for MIS
    product_order = [
        "V1", "V2", "V3", "V4", "V6", "V9", "V10",
        "V1- P of 2", "V2- P of 2", "V4- P of 2", "V4- P of 3",
        "V6- P of 2", "V9 P of 2",
        "V6-V1 Combo", "V6-V2 Combo",
        "V1-V2 Combo", "V1-V4 Combo", "V2-V4 Combo",
        "V9-V2 Combo", "V9-V3 Combo", "V9-V10 Combo",
        "Busy Book Blue", "Busy Book Pink", "Human Book",
        "Ganesha", "Krishna", "Hanuman",
        "Car", "Tank", "JCB",
    ]

    # COGS per unit (from previous MIS — user fills these)
    # Placeholder: 0 means user needs to fill
    cogs_map = {}

    row = 2
    for product in product_order:
        data = product_data.get(product, {
            "total_orders": 0, "shipped": 0, "delivered": 0,
            "rto": 0, "in_transit": 0, "cancelled": 0,
            "revenue": 0.0, "freight": 0.0
        })

        if data["total_orders"] == 0 and product not in product_data:
            continue  # Skip products with no orders

        r = row
        cogs_unit = cogs_map.get(product, 0)

        # A: Product name
        ws.cell(row=r, column=1, value=product).border = border
        # B: Delivered Revenue
        ws.cell(row=r, column=2, value=round(data["revenue"], 2)).border = border
        ws.cell(row=r, column=2).number_format = currency_fmt
        # C: Total Expense = COGS + Shipping (=I+K)
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=3, value=f"=I{r}+K{r}")
        ws.cell(row=r, column=3).number_format = currency_fmt
        # D: P/L = Revenue - Expense (=B-C)
        ws.cell(row=r, column=4).border = border
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
        ws.cell(row=r, column=4).number_format = currency_fmt
        # E: Profit % (=D/B)
        ws.cell(row=r, column=5).border = border
        ws.cell(row=r, column=5, value=f'=IF(B{r}=0,"",D{r}/B{r})')
        ws.cell(row=r, column=5).number_format = pct_fmt
        # F: P/pcs (=D/J)
        ws.cell(row=r, column=6).border = border
        ws.cell(row=r, column=6, value=f'=IF(J{r}=0,"",D{r}/J{r})')
        ws.cell(row=r, column=6).number_format = currency_fmt
        # G: Total Orders
        ws.cell(row=r, column=7, value=data["total_orders"]).border = border
        # H: Shipped
        ws.cell(row=r, column=8, value=data["shipped"]).border = border
        # I: Total COGS = COGS/Unit * Shipped (=R*H)
        ws.cell(row=r, column=9).border = border
        ws.cell(row=r, column=9, value=f"=R{r}*H{r}")
        ws.cell(row=r, column=9).number_format = currency_fmt
        # J: Delivered
        ws.cell(row=r, column=10, value=data["delivered"]).border = border
        # K: Shipping Charges (freight)
        ws.cell(row=r, column=11, value=round(data["freight"], 2)).border = border
        ws.cell(row=r, column=11).number_format = currency_fmt
        # L: RTO
        ws.cell(row=r, column=12, value=data["rto"]).border = border
        # M: In-Transit
        ws.cell(row=r, column=13, value=data["in_transit"]).border = border
        # N: RTO% = RTO/Shipped
        ws.cell(row=r, column=14).border = border
        ws.cell(row=r, column=14, value=f'=IF(H{r}=0,"",L{r}/H{r})')
        ws.cell(row=r, column=14).number_format = pct_fmt
        # O: Shipped% = Shipped/Total
        ws.cell(row=r, column=15).border = border
        ws.cell(row=r, column=15, value=f'=IF(G{r}=0,"",H{r}/G{r})')
        ws.cell(row=r, column=15).number_format = pct_fmt
        # P: Delivered% = Delivered/Total
        ws.cell(row=r, column=16).border = border
        ws.cell(row=r, column=16, value=f'=IF(G{r}=0,"",J{r}/G{r})')
        ws.cell(row=r, column=16).number_format = pct_fmt
        # Q: Cancellation% = (Total-Shipped)/Total
        ws.cell(row=r, column=17).border = border
        ws.cell(row=r, column=17, value=f'=IF(G{r}=0,"",(G{r}-H{r})/G{r})')
        ws.cell(row=r, column=17).number_format = pct_fmt
        # R: COGS/Unit (user fills)
        ws.cell(row=r, column=18, value=cogs_unit).border = border
        ws.cell(row=r, column=18).number_format = currency_fmt

        row += 1

    # Total row
    total_row = row
    last_data_row = row - 1
    ws.cell(row=total_row, column=1, value="TOTAL").border = border
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    for col in range(1, 19):
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).border = border
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    # SUM for numeric columns
    for col_letter, col_idx in [("B", 2), ("G", 7), ("H", 8), ("I", 9), ("J", 10),
                                 ("K", 11), ("L", 12), ("M", 13)]:
        ws.cell(row=total_row, column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{last_data_row})")
        ws.cell(row=total_row, column=col_idx).number_format = currency_fmt

    # Total row formulas
    t = total_row
    ws.cell(row=t, column=3, value=f"=I{t}+K{t}")  # Expense
    ws.cell(row=t, column=3).number_format = currency_fmt
    ws.cell(row=t, column=4, value=f"=B{t}-C{t}")  # P/L
    ws.cell(row=t, column=4).number_format = currency_fmt
    ws.cell(row=t, column=5, value=f'=IF(B{t}=0,"",D{t}/B{t})')  # Profit%
    ws.cell(row=t, column=5).number_format = pct_fmt
    ws.cell(row=t, column=6, value=f'=IF(J{t}=0,"",D{t}/J{t})')  # P/pcs
    ws.cell(row=t, column=6).number_format = currency_fmt
    ws.cell(row=t, column=14, value=f'=IF(H{t}=0,"",L{t}/H{t})')  # RTO%
    ws.cell(row=t, column=14).number_format = pct_fmt
    ws.cell(row=t, column=15, value=f'=IF(G{t}=0,"",H{t}/G{t})')  # Shipped%
    ws.cell(row=t, column=15).number_format = pct_fmt
    ws.cell(row=t, column=16, value=f'=IF(G{t}=0,"",J{t}/G{t})')  # Delivered%
    ws.cell(row=t, column=16).number_format = pct_fmt
    ws.cell(row=t, column=17, value=f'=IF(G{t}=0,"",(G{t}-H{t})/G{t})')  # Cancel%
    ws.cell(row=t, column=17).number_format = pct_fmt

    # Column widths
    col_widths = [18, 20, 16, 14, 10, 10, 12, 10, 14, 10, 16, 8, 11, 8, 10, 11, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "B2"

    output_path = os.path.join(BASE_DIR, "January 2026 MIS.xlsx")
    wb.save(output_path)
    print(f"\nMIS saved to: {output_path}")
    return output_path


# ── Main ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  JANUARY 2026 MIS GENERATOR")
    print("=" * 60)

    # Step 1: Fetch orders (or load from cache)
    raw_path = os.path.join(BASE_DIR, "jan_orders_raw.json")
    if os.path.exists(raw_path):
        print("Loading cached January orders from jan_orders_raw.json...")
        with open(raw_path) as f:
            orders = json.load(f)
        print(f"Loaded {len(orders)} orders from cache")
    else:
        orders = fetch_january_orders()
        if orders is None:
            print("Failed to fetch orders. Exiting.")
            exit(1)
        with open(raw_path, "w") as f:
            json.dump(orders, f, default=str)
        print(f"Raw orders saved to: {raw_path}")

    # Step 2: Load freight
    freight_map = load_freight_data()

    # Step 3: Process orders
    product_data = process_orders(orders, freight_map)

    # Save processed data
    processed_path = os.path.join(BASE_DIR, "jan_mis_data.json")
    with open(processed_path, "w") as f:
        json.dump(product_data, f, indent=2, default=str)
    print(f"Processed data saved to: {processed_path}")

    # Save freight data
    freight_path = os.path.join(BASE_DIR, "jan_freight_data.json")
    freight_by_product = {k: v["freight"] for k, v in product_data.items()}
    with open(freight_path, "w") as f:
        json.dump(freight_by_product, f, indent=2)
    print(f"Freight data saved to: {freight_path}")

    # Step 4: Generate Excel
    generate_mis_excel(product_data)

    # Summary
    print("\n" + "=" * 60)
    print("  JANUARY 2026 MIS SUMMARY")
    print("=" * 60)
    total_orders = sum(d["total_orders"] for d in product_data.values())
    total_delivered = sum(d["delivered"] for d in product_data.values())
    total_revenue = sum(d["revenue"] for d in product_data.values())
    total_rto = sum(d["rto"] for d in product_data.values())
    total_shipped = sum(d["shipped"] for d in product_data.values())
    total_freight = sum(d["freight"] for d in product_data.values())

    print(f"  Total Orders:     {total_orders:,}")
    print(f"  Total Shipped:    {total_shipped:,}")
    print(f"  Total Delivered:  {total_delivered:,}")
    print(f"  Total RTO:        {total_rto:,}")
    print(f"  Total Revenue:    ₹{total_revenue:,.0f}")
    print(f"  Total Freight:    ₹{total_freight:,.0f}")
    if total_shipped > 0:
        print(f"  RTO%:             {total_rto/total_shipped*100:.1f}%")
    if total_orders > 0:
        print(f"  Delivered%:       {total_delivered/total_orders*100:.1f}%")
    print("=" * 60)
    print("\nNOTE: COGS/Unit column (R) is set to 0 — please fill in manually.")
    print("Once COGS is filled, Total Expense, P/L, Profit%, and P/pcs will auto-calculate.")

#!/usr/bin/env python3
"""
Amazon MIS Generator — Oct 2025 to Feb 2026 (from scratch)
Processes: Order files, Settlement statements, Ad Spend reports + invoices
Outputs: Per-product monthly MIS JSON files

Columns: Products, Revenue, Orders, Delivered, COGS, COGS/Unit,
         Commission, FBA Fees, Closing Fee, Promos, Refund Amt,
         Total Amazon Fees, Ad Spend, Profit, Profit %
"""

import csv, json, os, re
from collections import defaultdict

BASE = "/Users/yashthombare/Desktop/Gravity/Shiprocket D2C data"

# ── SKU → Product Name mapping ──────────────────────────────────────────────
SKU_MAP = {
    "PortablebusyboardV1.5": "V1",
    "PortablebusyboardV01.5": "V1",
    "NEW_V1": "V1",
    "PortableBusyBoardV2": "V2",
    "PortablebusyboardV2": "V2",
    "NEW_V2": "V2",
    "QK-GUA5-RIKR": "V2",
    "bb_v03": "V3",
    "bb_v03 SF": "V3",
    "PortablebusyboardV03new": "V3",
    "PortablebusyboardV7": "V4",
    "PortableBusyboardV7": "V4",
    "PortablebusyboardV7 SF": "V4",
    "PortablebusyboardV6": "V6",
    "PortablebusyboardV6 SF": "V6",
    "PortableBusyBoard_V09": "V9",
    "new_PortableBusyBoard V9": "V9",
    "bb_v10": "V10",
    "bb_v10_SF": "V10",
    "bb_v14_SF": "V10",
    "busybook_blue": "Busy Book Blue",
    "Busybook01": "Busy Book Blue",
    "busybook_pink": "Busy Book Pink",
    "Busybookpink": "Busy Book Pink",
    "Humanbody01": "Human Book",
    "V1pack2": "V1- P of 2",
    "V1pack3": "V1- P of 3",
    "V4pack2": "V4- P of 2",
    "V4pack3": "V4- P of 3",
    "new_ComboV2_V4": "V2-V4 Combo",
    "new_ComboV2_V6": "V6-V2 Combo",
    "ComboV1_V6": "V6-V1 Combo",
    "ComboV1_V2": "V1-V2 Combo",
    "comboV1_V4": "V1-V4 Combo",
    "ComboV1_V4": "V1-V4 Combo",
    "Combo (V9/V10)": "V9-V10 Combo",
    "Ganesha_02": "Ganesha",
    "DIY_Tank01": "Tank",
    "CS Basics 1": "Drawing Board",
    "Busybookblue": "Busy Book Blue",
    "PortablebusyboardV5": "V5",
    "PortablebusyboardV05 SF": "V5",
    "PortablebusyboardV07": "V7 Police Cruiser",
    "PortablebusyboardV3new": "V3",
}

COGS_MAP = {
    "V1": 225, "V2": 275, "V3": 662, "V4": 170,
    "V1- P of 2": 531, "V1- P of 3": 531, "V2- P of 2": 649,
    "V4- P of 2": 401, "V4- P of 3": 368,
    "V6": 275, "V6- P of 2": 649, "V9": 778, "V9 P of 2": 1664, "V10": 1009,
    "Busy Book Pink": 300, "Busy Book Blue": 300, "Human Book": 300,
    "V9-V3 Combo": 1440, "V9-V10 Combo": 1787,
    "V1-V4 Combo": 404, "V6-V2 Combo": 612, "V1-V2 Combo": 524,
    "V2-V4 Combo": 488, "V9-V2 Combo": 488, "V6-V1 Combo": 608,
    "Ganesha": 290, "Krishna": 290, "Hanuman": 290,
    "Car": 540, "Tank": 862, "Drawing Board": 250, "JCB": 540,
    "V5": 225, "V7 Police Cruiser": 600, "V8": 700,
}

MONTHS_IN_SCOPE = ["Oct 2025", "Nov 2025", "Dec 2025", "Jan 2026", "Feb 2026"]

MONTH_NUM_MAP = {
    (2025, 10): "Oct 2025", (2025, 11): "Nov 2025",
    (2025, 12): "Dec 2025", (2026, 1): "Jan 2026", (2026, 2): "Feb 2026",
}


def map_sku(sku):
    if not sku:
        return None
    return SKU_MAP.get(sku.strip())


def month_from_iso(date_str):
    """Extract month key from '2025-10-31T18:28:30+00:00'."""
    if not date_str or len(date_str) < 7:
        return None
    try:
        parts = date_str[:10].split("-")
        return MONTH_NUM_MAP.get((int(parts[0]), int(parts[1])))
    except Exception:
        return None


def month_from_settlement(posted_date):
    """Extract month from '28.09.2025' or '07.12.2025 07:08:30 UTC'."""
    if not posted_date or len(posted_date) < 10:
        return None
    try:
        parts = posted_date.strip().split(".")
        day, month = int(parts[0]), int(parts[1])
        year = int(parts[2][:4])
        return MONTH_NUM_MAP.get((year, month))
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1: Parse Order Files → Revenue, Orders, Delivered
# ══════════════════════════════════════════════════════════════════════════════

def process_orders():
    print("=" * 60)
    print("  STEP 1: Processing Order Files (Fulfilments/)")
    print("=" * 60)

    data = {m: defaultdict(lambda: {
        "total_orders": 0, "delivered": 0, "cancelled": 0,
        "revenue": 0.0,
    }) for m in MONTHS_IN_SCOPE}

    unmapped_skus = defaultdict(int)
    fulfilment_dir = os.path.join(BASE, "Fulfilments")

    for fn in sorted(os.listdir(fulfilment_dir)):
        if not fn.endswith(".txt"):
            continue
        filepath = os.path.join(fulfilment_dir, fn)
        with open(filepath, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter="\t")
            for row in reader:
                m = month_from_iso(row.get("purchase-date", ""))
                if not m or m not in data:
                    continue

                sku = (row.get("sku") or "").strip()
                product = map_sku(sku)
                if not product:
                    if sku:
                        unmapped_skus[sku] += 1
                    continue

                status = (row.get("item-status") or row.get("order-status") or "").strip()
                try:
                    item_price = float(row.get("item-price") or 0)
                except Exception:
                    item_price = 0

                pd = data[m][product]
                pd["total_orders"] += 1

                if status == "Cancelled":
                    pd["cancelled"] += 1
                elif status in ("Shipped", "Shipped - Delivered to Buyer"):
                    pd["delivered"] += 1
                    pd["revenue"] += item_price

    # Print summary
    for m in MONTHS_IN_SCOPE:
        t_orders = sum(v["total_orders"] for v in data[m].values())
        t_del = sum(v["delivered"] for v in data[m].values())
        t_rev = sum(v["revenue"] for v in data[m].values())
        print(f"  {m}: {t_orders} orders, {t_del} delivered, ₹{t_rev/100000:.2f}L revenue")

    if unmapped_skus:
        print(f"\n  ⚠️  Unmapped SKUs ({len(unmapped_skus)} unique):")
        for sku, cnt in sorted(unmapped_skus.items(), key=lambda x: -x[1])[:10]:
            print(f"    {sku}: {cnt} rows")

    return data


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2: Parse Settlement Statements → Fee Breakdown per product
# ══════════════════════════════════════════════════════════════════════════════

def process_settlements():
    print("\n" + "=" * 60)
    print("  STEP 2: Processing Settlement Statements (Statments/)")
    print("=" * 60)

    fees = {m: defaultdict(lambda: {
        "commission": 0.0,
        "fba_fees": 0.0,
        "closing_fee": 0.0,
        "promos": 0.0,
        "refund_amt": 0.0,
    }) for m in MONTHS_IN_SCOPE}

    stmt_dir = os.path.join(BASE, "Statments")
    stmt_files = [os.path.join(stmt_dir, fn) for fn in os.listdir(stmt_dir) if fn.endswith(".txt")]

    for filepath in stmt_files:
        with open(filepath, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter="\t")
            for row in reader:
                posted = row.get("posted-date", "")
                m = month_from_settlement(posted)
                if not m or m not in fees:
                    continue

                sku = (row.get("sku") or "").strip()
                product = map_sku(sku)
                if not product:
                    continue

                tx_type = (row.get("transaction-type") or "").strip()
                desc = (row.get("amount-description") or "").strip()

                try:
                    amount = float(row.get("amount") or 0)
                except Exception:
                    continue

                pd = fees[m][product]

                if tx_type == "Order":
                    # Fees are negative (charged to seller)
                    if "Commission" in desc:
                        pd["commission"] += amount
                    elif "FBA" in desc:
                        pd["fba_fees"] += amount
                    elif "closing fee" in desc.lower():
                        pd["closing_fee"] += amount
                    elif desc == "Promo rebates":
                        pd["promos"] += amount

                elif tx_type == "Refund":
                    # Refund principal = money returned to buyer (negative)
                    if desc == "Principal":
                        pd["refund_amt"] += amount
                    # Fee reversals on refunds go back to respective buckets
                    elif "Commission" in desc:
                        pd["commission"] += amount
                    elif "FBA" in desc:
                        pd["fba_fees"] += amount
                    elif "closing fee" in desc.lower():
                        pd["closing_fee"] += amount

    # Print summary
    for m in MONTHS_IN_SCOPE:
        t_comm = sum(abs(v["commission"]) for v in fees[m].values())
        t_fba = sum(abs(v["fba_fees"]) for v in fees[m].values())
        t_close = sum(abs(v["closing_fee"]) for v in fees[m].values())
        t_promo = sum(abs(v["promos"]) for v in fees[m].values())
        t_refund = sum(abs(v["refund_amt"]) for v in fees[m].values())
        total = t_comm + t_fba + t_close + t_promo + t_refund
        print(f"  {m}: Comm ₹{t_comm/100000:.2f}L, FBA ₹{t_fba/100000:.2f}L, "
              f"Close ₹{t_close/100000:.2f}L, Promo ₹{t_promo/100000:.2f}L, "
              f"Refund ₹{t_refund/100000:.2f}L → Total ₹{total/100000:.2f}L")

    return fees


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3: Parse Ad Spend (Invoice + SP/SB/SD reports)
# ══════════════════════════════════════════════════════════════════════════════

def process_adspend(orders):
    print("\n" + "=" * 60)
    print("  STEP 3: Processing Ad Spend")
    print("=" * 60)

    # ── 3a: Monthly totals from ads invoice (inc GST = Amount paid) ──
    monthly_totals = {m: 0.0 for m in MONTHS_IN_SCOPE}
    mon_name_map = {
        "January": 1, "February": 2, "March": 3, "April": 4,
        "May": 5, "June": 6, "July": 7, "August": 8,
        "September": 9, "October": 10, "November": 11, "December": 12,
    }

    invoice_path = os.path.join(BASE, "ADS", "statement 20250312 to 20260312 (1).csv")
    with open(invoice_path, encoding="utf-8-sig") as f:
        # Skip note line if present
        first_line = f.readline()
        if "Country" not in first_line:
            # First line was a note, continue with reader
            reader = csv.DictReader(f)
        else:
            # First line was the header
            f.seek(0)
            reader = csv.DictReader(f)

        for row in reader:
            date_str = row.get("Invoice issue Date", "")
            m_match = re.match(r"(\d+)\s+(\w+),\s*(\d+)", date_str)
            if not m_match:
                continue
            day, mon_name, year = m_match.groups()
            mon_num = mon_name_map.get(mon_name, 0)
            month_key = MONTH_NUM_MAP.get((int(year), mon_num))
            if not month_key:
                continue

            # Use "Amount paid" (inc GST) as actual cost
            amt_str = (row.get("Amount paid (not converted)") or "")
            amt_str = amt_str.replace("₹", "").replace("\u20b9", "").replace(",", "").replace('"', "").strip()
            if not amt_str:
                # Fallback: Amount billed + Tax
                billed = (row.get("Amount billed (not converted)") or "")
                billed = billed.replace("₹", "").replace("\u20b9", "").replace(",", "").replace('"', "").strip()
                tax = (row.get("Tax amount billed (not converted)") or "")
                tax = tax.replace("₹", "").replace("\u20b9", "").replace(",", "").replace('"', "").strip()
                try:
                    amt = float(billed or 0) + float(tax or 0)
                except Exception:
                    continue
            else:
                try:
                    amt = float(amt_str)
                except Exception:
                    continue

            monthly_totals[month_key] += amt

    print("  Monthly ad spend from invoices (inc GST):")
    for m in MONTHS_IN_SCOPE:
        print(f"    {m}: ₹{monthly_totals[m]:>10,.0f} ({monthly_totals[m]/100000:.2f}L)")

    # ── 3b: Per-SKU data from Sponsored Products report (Jan + Feb) ──
    sp_per_product = {m: defaultdict(float) for m in MONTHS_IN_SCOPE}
    sp_monthly_total = {m: 0.0 for m in MONTHS_IN_SCOPE}

    xlsx_path = os.path.join(BASE, "ADS", "Sponsored_Products_Advertised_product_report.xlsx")
    if os.path.exists(xlsx_path):
        import openpyxl
        from datetime import datetime
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            start, end = row[0], row[1]
            sku = row[7]
            spend = float(row[13] or 0)
            if not start or not end or spend == 0:
                continue
            product = map_sku(str(sku).strip())
            if not product:
                continue

            total_days = (end - start).days + 1
            if total_days <= 0:
                continue

            # Pro-rate for each month
            for m_key, (y, mo) in [("Jan 2026", (2026, 1)), ("Feb 2026", (2026, 2))]:
                from calendar import monthrange
                _, last_day = monthrange(y, mo)
                m_start = max(start, datetime(y, mo, 1))
                m_end = min(end, datetime(y, mo, last_day))
                if m_end >= m_start:
                    m_days = (m_end - m_start).days + 1
                    ratio = m_days / total_days
                    allocated = spend * ratio
                    sp_per_product[m_key][product] += allocated
                    sp_monthly_total[m_key] += allocated

        for m in ["Jan 2026", "Feb 2026"]:
            if sp_monthly_total[m] > 0:
                print(f"\n  {m} SP per-SKU total (ex-GST from report): ₹{sp_monthly_total[m]:,.0f}")

    # ── 3c: SB + SD campaign totals per month ──
    sb_sd_total = {m: 0.0 for m in MONTHS_IN_SCOPE}

    for report_name in ["Sponsored_Brands_Campaign_report.xlsx", "Sponsored_Display_Campaign_report.xlsx"]:
        rpath = os.path.join(BASE, "ADS", report_name)
        if not os.path.exists(rpath):
            continue
        import openpyxl
        from datetime import datetime
        wb = openpyxl.load_workbook(rpath)
        ws = wb.active
        # Find "Spend" column index
        headers = [str(c.value or "").strip() for c in ws[1]]
        spend_idx = None
        for i, h in enumerate(headers):
            if h == "Spend":
                spend_idx = i
                break
        if spend_idx is None:
            print(f"  ⚠️  No 'Spend' column in {report_name}")
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            start, end = row[0], row[1]
            spend = float(row[spend_idx] or 0)
            if not start or not end or spend == 0:
                continue

            total_days = (end - start).days + 1
            if total_days <= 0:
                continue

            for m_key, (y, mo) in [("Jan 2026", (2026, 1)), ("Feb 2026", (2026, 2))]:
                from calendar import monthrange
                _, last_day = monthrange(y, mo)
                from datetime import datetime as dt
                m_start = max(start, dt(y, mo, 1))
                m_end = min(end, dt(y, mo, last_day))
                if m_end >= m_start:
                    m_days = (m_end - m_start).days + 1
                    ratio = m_days / total_days
                    sb_sd_total[m_key] += spend * ratio

        label = "SB" if "Brands" in report_name else "SD"
        for m in ["Jan 2026", "Feb 2026"]:
            if sb_sd_total[m] > 0:
                print(f"  {m} {label} spend (ex-GST): ₹{sb_sd_total[m]:,.0f}")

    # ── 3d: Build per-product ad spend for all months ──
    adspend_per_product = {m: {} for m in MONTHS_IN_SCOPE}

    for m in MONTHS_IN_SCOPE:
        total_ad = monthly_totals.get(m, 0)  # inc GST from invoice
        if total_ad == 0:
            continue

        if m in ("Jan 2026", "Feb 2026") and sp_monthly_total.get(m, 0) > 0:
            # Use SP report proportions, scale to invoice total
            # SP report is ex-GST, invoice is inc GST
            # Distribute invoice total using SP proportions for SP portion
            # SB+SD gets distributed by revenue

            sp_total_exgst = sp_monthly_total[m]
            sbsd_exgst = sb_sd_total.get(m, 0)
            all_reports_exgst = sp_total_exgst + sbsd_exgst

            # Scale factor: invoice total / reports total * 1.0
            # But invoice is inc GST, reports are ex-GST
            # Invoice ≈ (SP + SB + SD) × 1.18
            # So per-product ad spend (inc GST) = (SP spend / SP total) × invoice total

            for product, sp_spend in sp_per_product[m].items():
                # Product's share of total SP spend, applied to full invoice
                share = sp_spend / all_reports_exgst if all_reports_exgst > 0 else 0
                adspend_per_product[m][product] = round(total_ad * share, 2)

            # Any remainder (from SB/SD not mapped to products) → distribute by revenue
            allocated = sum(adspend_per_product[m].values())
            remainder = total_ad - allocated
            if remainder > 10:
                month_orders = orders.get(m, {})
                total_rev = sum(v["revenue"] for v in month_orders.values() if isinstance(v, dict))
                if total_rev > 0:
                    for product, v in month_orders.items():
                        if isinstance(v, dict) and v["revenue"] > 0:
                            share = v["revenue"] / total_rev
                            adspend_per_product[m][product] = adspend_per_product[m].get(product, 0) + round(remainder * share, 2)
        else:
            # Oct/Nov/Dec: distribute total by revenue proportion
            month_orders = orders.get(m, {})
            total_rev = sum(v["revenue"] for v in month_orders.values() if isinstance(v, dict))
            if total_rev == 0:
                continue
            for product, v in month_orders.items():
                if isinstance(v, dict) and v["revenue"] > 0:
                    share = v["revenue"] / total_rev
                    adspend_per_product[m][product] = round(total_ad * share, 2)

    print("\n  Per-product ad spend summary:")
    for m in MONTHS_IN_SCOPE:
        total = sum(adspend_per_product[m].values())
        print(f"    {m}: ₹{total:>10,.0f} ({total/100000:.2f}L) across {len(adspend_per_product[m])} products")

    return adspend_per_product, monthly_totals


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4: Build MIS JSON
# ══════════════════════════════════════════════════════════════════════════════

def build_mis(orders, fees, adspend_per_product, monthly_totals):
    print("\n" + "=" * 60)
    print("  STEP 4: Building MIS")
    print("=" * 60)

    all_data = {}

    for m in MONTHS_IN_SCOPE:
        month_data = {}
        all_products = set(orders[m].keys()) | set(fees[m].keys())

        for product in sorted(all_products):
            o = orders[m].get(product, {
                "total_orders": 0, "delivered": 0, "cancelled": 0, "revenue": 0.0,
            })
            f = fees[m].get(product, {
                "commission": 0.0, "fba_fees": 0.0, "closing_fee": 0.0,
                "promos": 0.0, "refund_amt": 0.0,
            })

            delivered = o["delivered"]
            cogs_unit = COGS_MAP.get(product, 0)
            cogs = cogs_unit * delivered
            revenue = round(o["revenue"], 2)

            commission = round(abs(f["commission"]), 2)
            fba_fees = round(abs(f["fba_fees"]), 2)
            closing_fee = round(abs(f["closing_fee"]), 2)
            promos = round(abs(f["promos"]), 2)
            refund_amt = round(abs(f["refund_amt"]), 2)
            total_amazon_fees = round(commission + fba_fees + closing_fee + promos, 2)
            net_revenue = round(revenue - refund_amt, 2)

            ad_spend = round(adspend_per_product.get(m, {}).get(product, 0), 2)
            profit = round(net_revenue - cogs - total_amazon_fees - ad_spend, 2)
            profit_pct = round(profit / net_revenue, 4) if net_revenue > 0 else 0

            if o["total_orders"] == 0 and revenue == 0 and total_amazon_fees == 0:
                continue

            month_data[product] = {
                "revenue": net_revenue,
                "total_orders": o["total_orders"],
                "delivered": delivered,
                "cogs": round(cogs, 2),
                "cogs_unit": cogs_unit,
                "commission": commission,
                "fba_fees": fba_fees,
                "closing_fee": closing_fee,
                "promos": promos,
                "refund_amt": refund_amt,
                "total_amazon_fees": total_amazon_fees,
                "ad_spend": ad_spend,
                "profit": profit,
                "profit_pct": profit_pct,
            }

        all_data[m] = month_data

        # Print summary
        t_orders = sum(v["total_orders"] for v in month_data.values())
        t_del = sum(v["delivered"] for v in month_data.values())
        t_rev = sum(v["revenue"] for v in month_data.values())
        t_cogs = sum(v["cogs"] for v in month_data.values())
        t_fees = sum(v["total_amazon_fees"] for v in month_data.values())
        t_ads = sum(v["ad_spend"] for v in month_data.values())
        t_profit = sum(v["profit"] for v in month_data.values())

        print(f"\n  {m}:")
        print(f"    Orders: {t_orders} | Delivered: {t_del}")
        print(f"    Revenue:          ₹{t_rev:>10,.0f} ({t_rev/100000:.2f}L)")
        print(f"    COGS:             ₹{t_cogs:>10,.0f} ({t_cogs/100000:.2f}L)")
        print(f"    Amazon Fees:      ₹{t_fees:>10,.0f} ({t_fees/100000:.2f}L)")
        print(f"    Ad Spend:         ₹{t_ads:>10,.0f} ({t_ads/100000:.2f}L)")
        print(f"    Profit:           ₹{t_profit:>10,.0f} ({t_profit/100000:.2f}L)")
        if t_rev > 0:
            print(f"    Profit %:         {t_profit/t_rev*100:.1f}%")

    # Save JSON files
    prefix_map = {
        "Oct 2025": "amazon_oct", "Nov 2025": "amazon_nov",
        "Dec 2025": "amazon_dec", "Jan 2026": "amazon_jan", "Feb 2026": "amazon_feb",
    }
    for m, d in all_data.items():
        path = os.path.join(BASE, f"{prefix_map[m]}_mis_data.json")
        with open(path, "w") as f:
            json.dump(d, f, indent=2)

    # Save monthly ad spend totals
    with open(os.path.join(BASE, "amazon_adspend_monthly.json"), "w") as f:
        json.dump(monthly_totals, f, indent=2)

    print("\n  JSON files saved ✓")
    return all_data


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("\n🔄 Amazon MIS Generator — Oct 2025 to Feb 2026\n")

    orders = process_orders()
    fees = process_settlements()
    adspend_per_product, monthly_totals = process_adspend(orders)
    mis = build_mis(orders, fees, adspend_per_product, monthly_totals)

    print("\n" + "=" * 60)
    print("  ✅ Amazon MIS Complete!")
    print("=" * 60)

#!/usr/bin/env python3
"""
Generate Instamart MIS data (Apr 2025 – Jan 2026).

Source: "For Instamart & Blinkit Reffrance.xlsx" (team's manual reference)
Instamart section in each month's sheet.

Revenue = what Swiggy pays us.
Units derived from: (Total Expense - COGS) / ₹20 per unit logistics.
COGS/unit = COGS / units (varies by month — Instamart pricing differs from D2C).
"""

import json, os
import openpyxl

BASE = "/Users/yashthombare/Desktop/Gravity/Shiprocket D2C data"
REF_FILE = os.path.join(BASE, "For Instamart & Blinkit Reffrance.xlsx")

SHEET_TO_MONTH = {
    "April 25": "Apr 2025", "May 25": "May 2025", "June 25": "Jun 2025",
    "July 25": "Jul 2025", "Aug 25": "Aug 2025", "Sep 25": "Sep 2025",
    "Oct 25": "Oct 2025", "Nov 25": "Nov 2025", "Dec 25": "Dec 2025",
    "Jan26": "Jan 2026",
}

LOGISTICS_PER_UNIT = 20  # ₹20/unit logistics as per reference

# Standard COGS/unit (same as D2C)
COGS_MAP = {"V1": 225, "V2": 275, "V4": 170}


def safe_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("₹", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def read_instamart_data():
    """Read Instamart sections from reference file."""
    wb = openpyxl.load_workbook(REF_FILE, data_only=True)
    monthly = {}

    for sheet_name, month_label in SHEET_TO_MONTH.items():
        ws = wb[sheet_name]

        # Find Instamart section
        insta_start = None
        for r in range(1, ws.max_row + 1):
            val2 = str(ws.cell(r, 2).value or "").strip().lower()
            if val2 == "instamart":
                insta_start = r
                break
        if not insta_start:
            continue

        # Data starts 3 rows after section header (title, blank, header)
        data_start = insta_start + 3

        # Read ad spend from summary below products
        ad_spend = 0
        for r in range(data_start, data_start + 20):
            label = str(ws.cell(r, 1).value or "").strip()
            if label == "Ad Spent":
                ad_spend = safe_float(ws.cell(r, 2).value)
                break

        products = {}
        for r in range(data_start, data_start + 10):
            name = str(ws.cell(r, 1).value or "").strip()
            if not name:
                break

            rev = safe_float(ws.cell(r, 2).value)
            expense = safe_float(ws.cell(r, 3).value)
            cogs = safe_float(ws.cell(r, 7).value)  # Product exp column

            # Derive units: (expense - COGS) = logistics = units × ₹20
            logistics = expense - cogs
            units = round(logistics / LOGISTICS_PER_UNIT) if logistics > 0 else 0
            cogs_unit = COGS_MAP.get(name, 0)

            if rev == 0 and units == 0:
                continue

            products[name] = {
                "revenue": round(rev, 2),
                "total_orders": units,
                "delivered": units,
                "returned": 0,
                "cancelled": 0,
                "cogs_unit": cogs_unit,
                "ad_spend": 0,  # allocated at month level below
            }

        # Add 18% GST to ad spend, then distribute proportionally by revenue
        ad_spend_with_gst = round(ad_spend * 1.18, 2)
        total_rev = sum(p["revenue"] for p in products.values())
        if ad_spend_with_gst > 0 and total_rev > 0:
            for p in products.values():
                p["ad_spend"] = round(ad_spend_with_gst * (p["revenue"] / total_rev), 2)

        monthly[month_label] = products

    wb.close()
    return monthly


def save_json_files(monthly):
    """Save per-month JSON files."""
    output_map = {
        "Apr 2025": "instamart_apr_2025_mis_data.json",
        "May 2025": "instamart_may_2025_mis_data.json",
        "Jun 2025": "instamart_jun_2025_mis_data.json",
        "Jul 2025": "instamart_jul_2025_mis_data.json",
        "Aug 2025": "instamart_aug_2025_mis_data.json",
        "Sep 2025": "instamart_sep_2025_mis_data.json",
        "Oct 2025": "instamart_oct_2025_mis_data.json",
        "Nov 2025": "instamart_nov_2025_mis_data.json",
        "Dec 2025": "instamart_dec_2025_mis_data.json",
        "Jan 2026": "instamart_jan_2026_mis_data.json",
    }

    for month_label, filename in output_map.items():
        products = monthly.get(month_label, {})
        output = {}
        for product, p in sorted(products.items()):
            if p["total_orders"] == 0 and p["revenue"] == 0:
                continue
            output[product] = {
                "revenue": round(p["revenue"], 2),
                "total_orders": p["total_orders"],
                "delivered": p["delivered"],
                "returned": p["returned"],
                "cancelled": p["cancelled"],
                "cogs_unit": p["cogs_unit"],
                "ad_spend": round(p["ad_spend"], 2),
            }

        filepath = os.path.join(BASE, filename)
        with open(filepath, "w") as f:
            json.dump(output, f, indent=2)

        total_orders = sum(p["total_orders"] for p in output.values())
        total_rev = sum(p["revenue"] for p in output.values())
        total_ads = sum(p["ad_spend"] for p in output.values())
        print(f"  {month_label}: {len(output)} products, "
              f"{total_orders} units, "
              f"₹{total_rev:,.0f} revenue, "
              f"₹{total_ads:,.0f} ad spend → {filename}")


def main():
    print("\nGenerating Instamart MIS (Apr 2025 – Jan 2026)\n")

    print("Reading reference file...")
    monthly = read_instamart_data()

    print("\nSaving JSON files...")
    save_json_files(monthly)

    # Summary
    print("\n" + "=" * 60)
    print("INSTAMART MIS SUMMARY")
    print("=" * 60)
    grand_orders = 0
    grand_revenue = 0
    for month in ["Apr 2025", "May 2025", "Jun 2025", "Jul 2025", "Aug 2025",
                   "Sep 2025", "Oct 2025", "Nov 2025", "Dec 2025", "Jan 2026"]:
        products = monthly.get(month, {})
        orders = sum(p["total_orders"] for p in products.values())
        rev = sum(p["revenue"] for p in products.values())
        grand_orders += orders
        grand_revenue += rev
        print(f"  {month:12s}: {orders:4d} units | ₹{rev:>10,.0f}")

    print(f"\n  TOTAL:        {grand_orders:4d} units | ₹{grand_revenue:>10,.0f}")
    print(f"\nDone!")


if __name__ == "__main__":
    main()

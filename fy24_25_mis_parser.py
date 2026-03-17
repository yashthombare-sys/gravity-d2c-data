#!/usr/bin/env python3
"""
Parse FY24-FY25 MIS.xlsx → generate D2C MIS JSON files for each month (Apr 2024 - Mar 2025).
Only extracts the Shopify/D2C section from each sheet.
Uses same COGS map as the current dashboard.
"""
import openpyxl
import json
import os

XLSX = 'FY24-FY25 MIS.xlsx'

# Sheet name → month label
SHEET_MAP = {
    'April24': 'Apr 2024', 'May': 'May 2024', 'June': 'Jun 2024',
    'July': 'Jul 2024', 'Aug': 'Aug 2024', 'Sept': 'Sep 2024',
    'Oct': 'Oct 2024', 'Nov': 'Nov 2024', 'Dec': 'Dec 2024',
    'Jan': 'Jan 2025', 'Feb 25': 'Feb 2025', 'March 25': 'Mar 2025'
}

# Product name normalization — map xlsx names to standard dashboard names
NAME_MAP = {
    'Busy Book': 'Busy Book Blue',
    'V4-V5 Combo': 'V4-V5 Combo',
    'V7-V4 Combo': 'V7-V4 Combo',
    'V6-V4 Combo': 'V6-V4 Combo',
}

# Same COGS_MAP as dashboard
COGS_MAP = {
    "V1": 225, "V2": 275, "V3": 662, "V4": 170, "V5": 225, "V6": 275, "V7": 600,
    "V7 Police Cruiser": 600, "V8": 700, "V9": 778, "V10": 1009,
    "V1- P of 2": 531, "V1- P of 3": 797, "V2- P of 2": 649,
    "V4- P of 2": 401, "V4- P of 3": 368, "V6- P of 2": 649,
    "V9 P of 2": 1664,
    "Busy Book Pink": 300, "Busy Book Blue": 300, "Human Book": 300,
    "V9-V3 Combo": 1440, "V9-V10 Combo": 1787,
    "V1-V4 Combo": 404, "V6-V2 Combo": 612, "V1-V2 Combo": 524,
    "V2-V4 Combo": 488, "V9-V2 Combo": 488, "V6-V1 Combo": 608,
    "Ganesha": 290, "Krishna": 290, "Hanuman": 290,
    "Car": 540, "Tank": 862, "JCB": 540,
    "CS Basics 1": 250, "Drawing Board": 250,
    "V1-Calculator Combo": 475, "Color Matching Game": 200,
    "V1-V6 Combo": 500, "V4-V6 Combo": 445, "V2-V9 Combo": 1053,
    # FY24-25 specific combos
    "V4-V5 Combo": 395, "V7-V4 Combo": 770, "V6-V4 Combo": 445,
    "V6- P of 3": 924,  # V6 × 3 + packing
    "Busy Book": 300,
}

SKIP_ROWS = {'Total', 'Total Revenue', 'Total Revenue ', 'Total Expance',
             'Net P&L', 'P&L', 'P&L %', 'COGS', 'Logistics', 'Marketing',
             'Shopify', 'Grand Total', 'TOTAL', 'Blended', 'AOV', 'CAC'}

def num(v):
    """Safe number conversion."""
    if v is None: return 0
    if isinstance(v, (int, float)): return v
    try: return float(v)
    except: return 0


def parse_sheet(ws, sheet_name):
    """Extract Shopify/D2C product rows from a sheet."""
    # Find header row (row with 'Products' in col A)
    header_row = None
    for r in range(1, min(15, ws.max_row + 1)):
        val = ws.cell(r, 1).value
        if val and str(val).strip() == 'Products':
            header_row = r
            break

    if not header_row:
        print(f"  WARNING: No header found in {sheet_name}")
        return {}

    # Read headers
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h and isinstance(h, str):
            headers[h.strip()] = c

    # Column mappings (handle varying column names)
    col_revenue = headers.get('Total revenue', headers.get('Total Revenue', None))
    col_orders = headers.get('Total orders', None)
    col_shipped = headers.get('Shipped', None)
    col_delivered = headers.get('Delivered', None)
    col_rto = headers.get('RTO', None)
    col_freight = headers.get('Shipping charges', None)
    col_adspent = headers.get('Ad spent', headers.get('Ad Spent', None))
    col_prodexp = headers.get('Product expense', None)
    col_intransit = headers.get('In-transit', headers.get('In-T Count', None))

    products = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or not isinstance(name, str):
            continue
        name = name.strip()

        # Stop at summary rows or section breaks
        if name in SKIP_ROWS or name.startswith('Total'):
            break

        # Skip empty/zero rows
        revenue = num(ws.cell(r, col_revenue).value) if col_revenue else 0
        orders = int(num(ws.cell(r, col_orders).value)) if col_orders else 0

        if revenue == 0 and orders == 0:
            continue

        # Normalize name
        prod_name = NAME_MAP.get(name, name)

        shipped = int(num(ws.cell(r, col_shipped).value)) if col_shipped else 0
        delivered = int(num(ws.cell(r, col_delivered).value)) if col_delivered else 0
        rto = int(num(ws.cell(r, col_rto).value)) if col_rto else 0
        freight = num(ws.cell(r, col_freight).value) if col_freight else 0
        adspent = num(ws.cell(r, col_adspent).value) if col_adspent else 0
        in_transit = int(num(ws.cell(r, col_intransit).value)) if col_intransit else 0

        cancelled = max(0, orders - shipped)

        products[prod_name] = {
            "total_orders": orders,
            "shipped": shipped,
            "delivered": delivered,
            "rto": rto,
            "in_transit": in_transit,
            "cancelled": cancelled,
            "lost": 0,
            "revenue": round(revenue, 2),
            "freight": round(freight, 2),
        }

    return products


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    all_data = {}

    for sheet_name, month_label in SHEET_MAP.items():
        ws = wb[sheet_name]
        print(f"Parsing {sheet_name} → {month_label}...")
        products = parse_sheet(ws, sheet_name)

        if not products:
            print(f"  WARNING: No products found!")
            continue

        all_data[month_label] = products

        # Summary
        total_rev = sum(p['revenue'] for p in products.values())
        total_orders = sum(p['total_orders'] for p in products.values())
        total_del = sum(p['delivered'] for p in products.values())
        total_rto = sum(p['rto'] for p in products.values())
        print(f"  {len(products)} products | {total_orders} orders | {total_del} delivered | {total_rto} RTO | ₹{total_rev/1e5:.2f}L revenue")

        # Flag products with no COGS mapping
        for pname in products:
            if pname not in COGS_MAP:
                print(f"  ⚠ No COGS mapping for: {pname}")

    # Save individual month files
    for month_label, products in all_data.items():
        # Convert month label to filename: "Apr 2024" → "apr_2024_mis_data.json"
        fname = month_label.lower().replace(' ', '_') + '_mis_data.json'
        with open(fname, 'w') as f:
            json.dump(products, f, indent=2)
        print(f"Saved: {fname}")

    # Print overall summary
    print(f"\n{'='*60}")
    print(f"FY24-25 D2C MIS Summary")
    print(f"{'='*60}")
    for month_label in SHEET_MAP.values():
        if month_label not in all_data:
            print(f"  {month_label}: MISSING")
            continue
        p = all_data[month_label]
        rev = sum(v['revenue'] for v in p.values())
        orders = sum(v['total_orders'] for v in p.values())
        delivered = sum(v['delivered'] for v in p.values())
        print(f"  {month_label}: {len(p)} products | {orders} orders | {delivered} delivered | ₹{rev/1e5:.2f}L")


if __name__ == '__main__':
    main()

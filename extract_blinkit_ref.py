#!/usr/bin/env python3
"""
Extract Blinkit section data from 'For Instamart & Blinkit Reffrance.xlsx'
Each month sheet has a Blinkit section with per-product data for V1, V4, V2, V6.
Blinkit naming: Toggle Play=V1, Mini Switch=V4, Spark Switch=V2, Tinker Pad=V6
"""

import json
import openpyxl

EXCEL_PATH = "/Users/yashthombare/Desktop/Gravity/Shiprocket D2C data/For Instamart & Blinkit Reffrance.xlsx"
OUTPUT_JSON = "/Users/yashthombare/Desktop/Gravity/Shiprocket D2C data/blinkit_ref_data.json"

BLINKIT_NAME_MAP = {
    "Toggle Play": "V1",
    "Mini Switch": "V4",
    "Spark Switch": "V2",
    "Tinker Pad": "V6",
}

# Column mapping for Blinkit section (consistent across all months):
# A=Products, B=Total revenue, C=Total expense, D=Total P/L, E=Profit %,
# F=P/pcs, G=Product exp, H=Total orders, I=Blinkit exp, J=Ads, K=Logistics

SHEET_NAMES = [
    "April 25", "May 25", "June 25", "July 25", "Aug 25",
    "Sep 25", "Oct 25", "Nov 25", "Dec 25", "Jan26"
]

MONTH_LABELS = [
    "Apr 2025", "May 2025", "Jun 2025", "Jul 2025", "Aug 2025",
    "Sep 2025", "Oct 2025", "Nov 2025", "Dec 2025", "Jan 2026"
]


def safe_num(val):
    """Convert value to number, return 0 for None/errors."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        # Handle Excel error strings
        if val.startswith("#") or val.strip() == "":
            return 0
        try:
            return float(val.replace(",", ""))
        except ValueError:
            return 0
    return 0


def find_blinkit_section(ws):
    """Find the Blinkit header row and extract product data rows + summary."""
    blinkit_header_row = None

    # Find the row where column B says "Blinkit"
    for row_idx in range(1, ws.max_row + 1):
        b_val = ws.cell(row_idx, 2).value  # Column B
        if b_val and isinstance(b_val, str) and "blinkit" in b_val.lower():
            blinkit_header_row = row_idx
            break

    if blinkit_header_row is None:
        return None, None, None

    # The column headers are 2 rows below the "Blinkit" label
    header_row = blinkit_header_row + 2

    # Verify header row has expected columns
    header_b = ws.cell(header_row, 2).value
    if not (header_b and isinstance(header_b, str) and "revenue" in header_b.lower()):
        # Try header_row = blinkit_header_row + 1
        header_row = blinkit_header_row + 1
        header_b = ws.cell(header_row, 2).value
        if not (header_b and isinstance(header_b, str) and "revenue" in header_b.lower()):
            return None, None, None

    # Read product rows (starting from header_row + 1)
    products = []
    data_start = header_row + 1
    total_row_data = None

    for r in range(data_start, data_start + 20):  # Max 20 rows to search
        a_val = ws.cell(r, 1).value
        b_val = ws.cell(r, 2).value

        # Product row: column A has a product name (V1, V2, V4, V6)
        if a_val and isinstance(a_val, str) and a_val.strip():
            product_name = a_val.strip()
            product = {
                "product": product_name,
                "total_revenue": safe_num(ws.cell(r, 2).value),
                "total_expense": safe_num(ws.cell(r, 3).value),
                "total_pl": safe_num(ws.cell(r, 4).value),
                "profit_pct": safe_num(ws.cell(r, 5).value),
                "p_per_pcs": safe_num(ws.cell(r, 6).value),
                "product_exp": safe_num(ws.cell(r, 7).value),
                "total_orders": safe_num(ws.cell(r, 8).value),
                "blinkit_exp": safe_num(ws.cell(r, 9).value),
                "ads": safe_num(ws.cell(r, 10).value),
                "logistics": safe_num(ws.cell(r, 11).value),
            }
            products.append(product)
        elif b_val is not None and a_val is None:
            # Total row (no product name in A, but has values in B onwards)
            total_row_data = {
                "total_revenue": safe_num(ws.cell(r, 2).value),
                "total_expense": safe_num(ws.cell(r, 3).value),
                "total_pl": safe_num(ws.cell(r, 4).value),
                "product_exp": safe_num(ws.cell(r, 7).value),
                "total_orders": safe_num(ws.cell(r, 8).value),
                "ads": safe_num(ws.cell(r, 10).value),
                "logistics": safe_num(ws.cell(r, 11).value),
            }
            break

    # Also grab the Blinkit summary section (below the product rows)
    # Stop if we hit the next channel section (Instamart, FirstCry, etc.)
    summary = {}
    for r in range(data_start, data_start + 30):
        a_val = ws.cell(r, 1).value
        b_val = ws.cell(r, 2).value
        # Stop at the next channel section header
        if b_val and isinstance(b_val, str) and any(
            x in b_val.lower() for x in ["instamart", "firstcry", "cred", "flipkart", "shopify", "amazon"]
        ):
            break
        if a_val and isinstance(a_val, str):
            key = a_val.strip().lower()
            # Stop if we hit a month label + channel name (next section)
            if any(m in key for m in ["april", "may", "june", "july", "aug", "sep", "oct", "nov", "dec", "jan", "feb"]):
                break
            if "ad" in key and "spent" in key:
                summary["ad_spent_total"] = safe_num(ws.cell(r, 2).value)
            elif key.startswith("total revenue"):
                summary["total_revenue"] = safe_num(ws.cell(r, 2).value)
            elif key == "cogs":
                summary["cogs"] = safe_num(ws.cell(r, 2).value)
            elif "logit" in key or "logis" in key:
                summary["logistics"] = safe_num(ws.cell(r, 2).value)
            elif "net p&l" in key:
                summary["net_pl"] = safe_num(ws.cell(r, 2).value)
            elif "p&l %" in key:
                summary["pl_pct"] = safe_num(ws.cell(r, 2).value)

    return products, total_row_data, summary


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    all_data = {}

    print("=" * 100)
    print("BLINKIT DATA EXTRACTION — For Instamart & Blinkit Reffrance.xlsx")
    print("Blinkit Products: V1 (Toggle Play), V4 (Mini Switch), V2 (Spark Switch), V6 (Tinker Pad)")
    print("=" * 100)

    for sheet_name, month_label in zip(SHEET_NAMES, MONTH_LABELS):
        ws = wb[sheet_name]
        products, total_row, summary = find_blinkit_section(ws)

        print(f"\n{'─' * 100}")
        print(f"  {month_label} (Sheet: {sheet_name})")
        print(f"{'─' * 100}")

        if products is None:
            print("  [!] Blinkit section NOT FOUND in this sheet")
            all_data[month_label] = {"error": "Blinkit section not found"}
            continue

        # Print product table
        print(f"  {'Product':<8} {'Revenue':>12} {'Expense':>12} {'P/L':>12} {'Profit%':>9} {'P/pcs':>8} {'Prod Exp':>12} {'Orders':>8} {'Blinkit Exp':>12} {'Ads':>12} {'Logistics':>10}")
        print(f"  {'─'*8} {'─'*12} {'─'*12} {'─'*12} {'─'*9} {'─'*8} {'─'*12} {'─'*8} {'─'*12} {'─'*12} {'─'*10}")

        for p in products:
            print(f"  {p['product']:<8} {p['total_revenue']:>12,.0f} {p['total_expense']:>12,.0f} {p['total_pl']:>12,.0f} {p['profit_pct']:>8.1f}% {p['p_per_pcs']:>8,.0f} {p['product_exp']:>12,.0f} {p['total_orders']:>8,.0f} {p['blinkit_exp']:>12,.0f} {p['ads']:>12,.0f} {p['logistics']:>10,.0f}")

        if total_row:
            print(f"  {'─'*8} {'─'*12} {'─'*12} {'─'*12} {'─'*9} {'─'*8} {'─'*12} {'─'*8} {'─'*12} {'─'*12} {'─'*10}")
            print(f"  {'TOTAL':<8} {total_row['total_revenue']:>12,.0f} {total_row['total_expense']:>12,.0f} {total_row['total_pl']:>12,.0f} {'':>9} {'':>8} {total_row['product_exp']:>12,.0f} {total_row['total_orders']:>8,.0f} {'':>12} {total_row['ads']:>12,.0f} {total_row['logistics']:>10,.0f}")

        if summary:
            print(f"\n  Summary:")
            if "total_revenue" in summary:
                print(f"    Total Revenue:  Rs.{summary['total_revenue']:>12,.0f}")
            if "cogs" in summary:
                print(f"    COGS:           Rs.{summary['cogs']:>12,.0f}")
            if "logistics" in summary:
                print(f"    Logistics:      Rs.{summary['logistics']:>12,.0f}")
            if "ad_spent_total" in summary:
                print(f"    Ad Spent:       Rs.{summary['ad_spent_total']:>12,.0f}")
            if "net_pl" in summary:
                print(f"    Net P&L:        Rs.{summary['net_pl']:>12,.0f}")
            if "pl_pct" in summary:
                print(f"    P&L %:          {summary['pl_pct']:>12.2f}%")

        # Store for JSON
        month_data = {
            "month": month_label,
            "products": products,
            "totals": total_row,
            "summary": summary,
        }
        all_data[month_label] = month_data

    # Grand summary across all months
    print(f"\n\n{'=' * 100}")
    print("GRAND SUMMARY — All Months")
    print(f"{'=' * 100}")
    print(f"  {'Month':<12} {'Revenue':>12} {'Expense':>12} {'P/L':>12} {'Orders':>8} {'Ads':>12} {'Logistics':>10} {'P&L %':>8}")
    print(f"  {'─'*12} {'─'*12} {'─'*12} {'─'*12} {'─'*8} {'─'*12} {'─'*10} {'─'*8}")

    grand_revenue = 0
    grand_expense = 0
    grand_pl = 0
    grand_orders = 0
    grand_ads = 0
    grand_logistics = 0

    for month_label in MONTH_LABELS:
        d = all_data.get(month_label, {})
        if "error" in d:
            print(f"  {month_label:<12} {'N/A':>12}")
            continue

        t = d.get("totals", {}) or {}
        s = d.get("summary", {}) or {}

        rev = t.get("total_revenue", 0) or s.get("total_revenue", 0)
        exp = t.get("total_expense", 0)
        pl = t.get("total_pl", 0) or s.get("net_pl", 0)
        orders = t.get("total_orders", 0)
        ads = s.get("ad_spent_total", 0) or t.get("ads", 0)
        logistics = t.get("logistics", 0) or s.get("logistics", 0)
        pl_pct = s.get("pl_pct", 0)

        grand_revenue += rev
        grand_expense += exp
        grand_pl += pl
        grand_orders += orders
        grand_ads += ads
        grand_logistics += logistics

        print(f"  {month_label:<12} {rev:>12,.0f} {exp:>12,.0f} {pl:>12,.0f} {orders:>8,.0f} {ads:>12,.0f} {logistics:>10,.0f} {pl_pct:>7.1f}%")

    print(f"  {'─'*12} {'─'*12} {'─'*12} {'─'*12} {'─'*8} {'─'*12} {'─'*10} {'─'*8}")
    grand_pl_pct = (grand_pl / grand_revenue * 100) if grand_revenue else 0
    print(f"  {'TOTAL':<12} {grand_revenue:>12,.0f} {grand_expense:>12,.0f} {grand_pl:>12,.0f} {grand_orders:>8,.0f} {grand_ads:>12,.0f} {grand_logistics:>10,.0f} {grand_pl_pct:>7.1f}%")

    # Per-product summary across all months
    print(f"\n\n{'=' * 100}")
    print("PER-PRODUCT SUMMARY — All Months Combined")
    print(f"{'=' * 100}")

    product_totals = {}
    for month_label in MONTH_LABELS:
        d = all_data.get(month_label, {})
        if "error" in d:
            continue
        for p in d.get("products", []):
            name = p["product"]
            if name not in product_totals:
                product_totals[name] = {"revenue": 0, "expense": 0, "pl": 0, "orders": 0, "product_exp": 0, "ads": 0, "logistics": 0}
            product_totals[name]["revenue"] += p["total_revenue"]
            product_totals[name]["expense"] += p["total_expense"]
            product_totals[name]["pl"] += p["total_pl"]
            product_totals[name]["orders"] += p["total_orders"]
            product_totals[name]["product_exp"] += p["product_exp"]
            product_totals[name]["ads"] += p["ads"]
            product_totals[name]["logistics"] += p["logistics"]

    print(f"  {'Product':<8} {'Revenue':>12} {'Expense':>12} {'P/L':>12} {'Orders':>8} {'Prod Exp':>12} {'Ads':>12} {'Logistics':>10} {'Profit%':>8}")
    print(f"  {'─'*8} {'─'*12} {'─'*12} {'─'*12} {'─'*8} {'─'*12} {'─'*12} {'─'*10} {'─'*8}")

    for name in ["V1", "V4", "V2", "V6"]:
        if name in product_totals:
            pt = product_totals[name]
            pct = (pt["pl"] / pt["revenue"] * 100) if pt["revenue"] else 0
            print(f"  {name:<8} {pt['revenue']:>12,.0f} {pt['expense']:>12,.0f} {pt['pl']:>12,.0f} {pt['orders']:>8,.0f} {pt['product_exp']:>12,.0f} {pt['ads']:>12,.0f} {pt['logistics']:>10,.0f} {pct:>7.1f}%")

    # Save to JSON
    with open(OUTPUT_JSON, "w") as f:
        json.dump(all_data, f, indent=2, default=str)

    print(f"\nData saved to: {OUTPUT_JSON}")


if __name__ == "__main__":
    main()

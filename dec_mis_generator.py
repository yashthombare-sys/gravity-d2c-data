#!/usr/bin/env python3
"""
December 2025 MIS Generator for Shiprocket D2C Data
Fetches orders via Shiprocket API, processes per MIS flow, saves JSON + pushes to Google Sheets.
"""

import requests
import json
import re
import os
import time
from collections import defaultdict
import gspread
from google.oauth2.service_account import Credentials
import openpyxl

# ── Config ──────────────────────────────────────────────────────────────────
BASE_DIR = "/Users/yashthombare/Desktop/Gravity/Shiprocket D2C data"
API_BASE = "https://apiv2.shiprocket.in/v1/external"
DEC_START = "2025-12-01"
DEC_END = "2025-12-31"
SHEET_URL = "https://docs.google.com/spreadsheets/d/1-aln640f4OxRmoS9R5EBvnQACp6edzxrMQDU6sgd3Lc/"
CREDS_FILE = f"{BASE_DIR}/shiproket-mis-70c28ae6e7fb.json"

# Load token from .env
with open(os.path.join(BASE_DIR, ".env")) as f:
    for line in f:
        if line.startswith("SHIPROCKET_API_TOKEN="):
            TOKEN = line.strip().split("=", 1)[1]

HEADERS_API = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {TOKEN}"
}

# ── Status Classification ───────────────────────────────────────────────────
DELIVERED_STATUSES = {"DELIVERED"}
RTO_STATUSES = {"RTO DELIVERED", "RTO IN TRANSIT", "RTO INITIATED", "RTO OFD",
                "REACHED BACK AT SELLER CITY", "REACHED BACK AT_SELLER_CITY"}
CANCELLED_STATUSES = {"CANCELED", "CANCELLATION REQUESTED"}
LOST_STATUSES = {"LOST", "DESTROYED", "MISROUTED", "UNTRACEABLE"}
IN_TRANSIT_STATUSES = {"IN TRANSIT", "IN TRANSIT-EN-ROUTE", "OUT FOR DELIVERY",
                       "PICKED UP", "OUT FOR PICKUP", "NEW ORDER",
                       "REACHED DESTINATION HUB", "UNDELIVERED-1ST ATTEMPT",
                       "UNDELIVERED-2ND ATTEMPT", "UNDELIVERED-3RD ATTEMPT"}
SKIP_STATUSES = {"SELF FULFILED", "QC FAILED", "RETURN DELIVERED",
                 "RETURN IN TRANSIT", "RETURN PENDING", "RETURN CANCELLED"}

SPARE_PARTS_KEYWORDS = [
    "RC TANK MOTOR", "RC CAR PCB", "Charging cable", "Documents",
    "charging cable", "spare", "SPARE", "pcb", "motor", "document"
]

# ── Product Mapping ─────────────────────────────────────────────────────────
PRODUCT_PATTERNS = [
    # Combos first
    (r"V9.*V10|V10.*V9|V9\+V10|V9-V10|V9 \+ V10|V9 V10 Combo", "V9-V10 Combo"),
    (r"V9.*V3|V3.*V9|V9\+V3|V9-V3|V9 \+ V3|V9 V3 Combo", "V9-V3 Combo"),
    (r"V9.*V2|V2.*V9|V9\+V2|V9-V2|V9 \+ V2|V9 V2 Combo", "V9-V2 Combo"),
    (r"V6.*V1|V1.*V6|V6\+V1|V6-V1|V6 \+ V1|V6 V1 Combo", "V6-V1 Combo"),
    (r"V6.*V2|V2.*V6|V6\+V2|V6-V2|V6 \+ V2|V6 V2 Combo", "V6-V2 Combo"),
    (r"V1.*V4|V4.*V1|V1\+V4|V1-V4|V1 \+ V4|V1 V4 Combo", "V1-V4 Combo"),
    (r"V1.*V2|V2.*V1|V1\+V2|V1-V2|V1 \+ V2|V1 V2 Combo", "V1-V2 Combo"),
    (r"V2.*V4|V4.*V2|V2\+V4|V2-V4|V2 \+ V4|V2 V4 Combo", "V2-V4 Combo"),
    # Pack of 2/3 variants
    (r"V9.*(?:Pack of 2|P of 2|pack of 2|2\s*pack)", "V9 P of 2"),
    (r"V6.*(?:Pack of 2|P of 2|pack of 2|2\s*pack)", "V6- P of 2"),
    (r"V4.*(?:Pack of 3|P of 3|pack of 3|3\s*pack)", "V4- P of 3"),
    (r"V4.*(?:Pack of 2|P of 2|pack of 2|2\s*pack)", "V4- P of 2"),
    (r"V2.*(?:Pack of 2|P of 2|pack of 2|2\s*pack)", "V2- P of 2"),
    (r"V1.*(?:Pack of 2|P of 2|pack of 2|2\s*pack)", "V1- P of 2"),
    # Individual products (check V10 before V1)
    (r"V10(?!\d)", "V10"),
    (r"V1(?!\d)", "V1"),
    (r"V2(?!\d)", "V2"),
    (r"V3(?!\d)", "V3"),
    (r"V4(?!\d)", "V4"),
    (r"V6(?!\d)", "V6"),
    (r"V9(?!\d)", "V9"),
    # Busy Books
    (r"(?i)busy\s*book.*(?:blue|boy)", "Busy Book Blue"),
    (r"(?i)busy\s*book.*(?:pink|girl)", "Busy Book Pink"),
    (r"(?i)human\s*(?:body\s*)?(?:busy\s*)?book", "Human Book"),
    # ClapCuddles
    (r"(?i)(?:clap\s*cuddle|clapcuddle).*(?:ganesh|ganesha)", "Ganesha"),
    (r"(?i)(?:clap\s*cuddle|clapcuddle).*(?:krishna)", "Krishna"),
    (r"(?i)(?:clap\s*cuddle|clapcuddle).*(?:hanuman)", "Hanuman"),
    (r"(?i)ganesh", "Ganesha"),
    (r"(?i)krishna", "Krishna"),
    (r"(?i)hanuman", "Hanuman"),
    # RC vehicles
    (r"(?i)(?:sooper\s*brains?\s*)?(?:rc\s*)?(?:army\s*)?tank", "Tank"),
    (r"(?i)(?:sooper\s*brains?\s*)?(?:rc\s*)?(?:racer|car)", "Car"),
    (r"(?i)(?:sooper\s*brains?\s*)?jcb", "JCB"),
]

COGS_MAP = {
    "V1": 225, "V2": 275, "V3": 662, "V4": 170,
    "V1- P of 2": 531, "V2- P of 2": 649, "V4- P of 2": 401, "V4- P of 3": 368,
    "V6": 275, "V6- P of 2": 649, "V9": 778, "V9 P of 2": 1664, "V10": 1009,
    "Busy Book Pink": 300, "Busy Book Blue": 300, "Human Book": 300,
    "V9-V3 Combo": 1440, "V9-V10 Combo": 1787,
    "V1-V4 Combo": 404, "V6-V2 Combo": 612, "V1-V2 Combo": 524,
    "V2-V4 Combo": 488, "V9-V2 Combo": 488, "V6-V1 Combo": 608,
    "Ganesha": 290, "Krishna": 290, "Hanuman": 290,
    "Car": 540, "Tank": 862,
}

# ── Categories (same as push_categorized_mis.py) ───────────────────────────
CATEGORIES = [
    {
        "name": "BUSY BOARD CATEGORY",
        "color": {"red": 0.933, "green": 0.522, "blue": 0.133},
        "products": [
            "V1", "V2", "V3", "V4", "V6", "V9", "V10",
            "V1- P of 2", "V2- P of 2", "V4- P of 2", "V4- P of 3",
            "V6- P of 2", "V9 P of 2",
            "V6-V1 Combo", "V6-V2 Combo",
            "V1-V2 Combo", "V1-V4 Combo", "V2-V4 Combo",
            "V9-V2 Combo", "V9-V3 Combo", "V9-V10 Combo",
            "Busy Book Blue", "Busy Book Pink", "Human Book",
        ],
    },
    {
        "name": "SOFT TOY CATEGORY",
        "color": {"red": 0.678, "green": 0.847, "blue": 0.902},
        "products": ["Ganesha", "Krishna", "Hanuman"],
    },
    {
        "name": "STEM CATEGORY",
        "color": {"red": 0.576, "green": 0.769, "blue": 0.490},
        "products": ["Car", "Tank", "JCB"],
    },
]

SHEET_HEADERS = [
    "Products", "Total Delivered Revenue", "Total Expense", "Total P/L",
    "Profit %", "P/pcs", "Total Orders", "Shipped", "Total COGS",
    "Delivered", "Shipping Charges", "RTO", "In-Transit",
    "RTO%", "Shipped%", "Delivered%", "Cancellation%", "COGS/Unit",
]


def classify_product(product_name):
    if not product_name:
        return None
    for pattern, category in PRODUCT_PATTERNS:
        if re.search(pattern, product_name):
            return category
    return None


def classify_status(status):
    if not status:
        return "unknown"
    s = status.upper().strip()
    if s in DELIVERED_STATUSES:
        return "delivered"
    if s in RTO_STATUSES:
        return "rto"
    if s in CANCELLED_STATUSES:
        return "cancelled"
    if s in SKIP_STATUSES or s.startswith("RETURN"):
        return "skip"
    if s in LOST_STATUSES:
        return "lost"
    if s in IN_TRANSIT_STATUSES:
        return "in_transit"
    return "skip"


def is_spare_part(product_name):
    if not product_name:
        return True
    name_lower = product_name.lower()
    for kw in SPARE_PARTS_KEYWORDS:
        if kw.lower() in name_lower:
            return True
    return False


# ── Step 1: Fetch orders ────────────────────────────────────────────────────
def fetch_december_orders():
    all_orders = []
    page = 1
    per_page = 200

    print("Fetching December 2025 orders from Shiprocket API...")

    while True:
        url = f"{API_BASE}/orders?per_page={per_page}&page={page}&from={DEC_START}&to={DEC_END}"
        print(f"  Page {page}...", end=" ", flush=True)

        resp = requests.get(url, headers=HEADERS_API)
        if resp.status_code == 401:
            print("\nERROR: Token expired. Re-authenticate and update .env")
            return None
        if resp.status_code != 200:
            print(f"\nERROR: API returned {resp.status_code}: {resp.text[:200]}")
            return None

        data = resp.json()

        if isinstance(data, dict) and "data" in data:
            orders = data["data"]
            if isinstance(orders, dict):
                orders = orders.get("data", orders.get("orders", []))
            if not isinstance(orders, list):
                orders = []
        elif isinstance(data, list):
            orders = data
        else:
            orders = []

        if not orders:
            print("0 orders (done)")
            break

        all_orders.extend(orders)
        print(f"{len(orders)} orders")

        if isinstance(data, dict) and "data" in data and isinstance(data["data"], dict):
            last_page = data["data"].get("last_page", 0)
            if last_page and page >= last_page:
                break

        meta = data.get("meta", {}) if isinstance(data, dict) else {}
        pagination = meta.get("pagination", {})
        total_pages = pagination.get("total_pages", 0)
        if total_pages and page >= total_pages:
            break

        if len(orders) < per_page:
            break

        page += 1

    print(f"\nTotal orders fetched: {len(all_orders)}")
    return all_orders


# ── Step 2: Load freight ────────────────────────────────────────────────────
def load_freight_data():
    freight_file = os.path.join(BASE_DIR, "Freight Total Amount.xlsx")
    wb = openpyxl.load_workbook(freight_file, read_only=True, data_only=True)
    ws = wb.active

    headers = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = col_idx

    order_col = None
    freight_col = None
    for name in ["order id", "orderid", "order_id", "sr order id"]:
        if name in headers:
            order_col = headers[name]
            break
    for name in ["freight total amount", "freight", "total freight", "freight total"]:
        if name in headers:
            freight_col = headers[name]
            break

    if not order_col or not freight_col:
        order_col = 1
        freight_col = 2

    freight_map = {}
    for row in ws.iter_rows(min_row=2):
        order_id = row[order_col - 1].value
        freight_val = row[freight_col - 1].value
        if order_id and freight_val:
            order_id = str(order_id).strip()
            if order_id not in freight_map:
                try:
                    freight_map[order_id] = float(freight_val)
                except (ValueError, TypeError):
                    pass

    wb.close()
    print(f"Loaded freight for {len(freight_map)} unique orders")
    return freight_map


# ── Step 3: Process orders ──────────────────────────────────────────────────
def process_orders(orders, freight_map):
    product_data = defaultdict(lambda: {
        "total_orders": 0, "shipped": 0, "delivered": 0,
        "rto": 0, "in_transit": 0, "cancelled": 0,
        "revenue": 0.0, "freight": 0.0,
    })

    seen = set()
    unmapped = defaultdict(int)
    skipped_statuses = 0
    spare_parts_skipped = 0

    for order in orders:
        order_id = str(order.get("id", order.get("order_id", "")))
        channel_order_id = str(order.get("channel_order_id", order_id))
        is_reverse = order.get("is_reverse", False)

        if is_reverse or str(is_reverse).lower() == "yes" or str(is_reverse) == "1":
            continue

        status_raw = order.get("status", order.get("status_code", ""))
        status = classify_status(status_raw)

        if status == "skip":
            skipped_statuses += 1
            continue

        products = order.get("products", order.get("order_items", []))
        if not products:
            product_name = order.get("product_name", "")
            if product_name:
                products = [{
                    "name": product_name,
                    "selling_price": order.get("product_price", order.get("selling_price", 0)),
                    "discount": order.get("discount", 0),
                    "quantity": order.get("product_quantity", order.get("quantity", 1)),
                }]

        order_line_values = []
        order_products_info = []

        for prod in products:
            pname = prod.get("name", prod.get("product_name", ""))
            price = float(prod.get("price", prod.get("selling_price", prod.get("product_price", 0))) or 0)
            discount = float(prod.get("discount", 0) or 0)
            qty = int(prod.get("quantity", prod.get("product_quantity", 1)) or 1)

            if is_spare_part(pname):
                spare_parts_skipped += 1
                continue
            if qty > 10:
                continue

            category = classify_product(pname)
            if not category:
                unmapped[pname] += 1
                continue

            dedup_key = (order_id, category)
            if dedup_key in seen:
                continue
            seen.add(dedup_key)

            line_value = (price - discount) * qty
            order_line_values.append(line_value)
            order_products_info.append((category, price, discount, qty, line_value))

        total_order_value = sum(order_line_values) if order_line_values else 0
        order_freight = freight_map.get(channel_order_id, freight_map.get(order_id, 0))

        for category, price, discount, qty, line_value in order_products_info:
            pd = product_data[category]
            pd["total_orders"] += 1

            if status == "delivered":
                pd["delivered"] += 1
                pd["shipped"] += 1
                pd["revenue"] += line_value
            elif status == "rto":
                pd["rto"] += 1
                pd["shipped"] += 1
            elif status == "in_transit":
                pd["in_transit"] += 1
                pd["shipped"] += 1
            elif status == "cancelled":
                pd["cancelled"] += 1

            if total_order_value > 0 and order_freight > 0:
                freight_share = (line_value / total_order_value) * order_freight
                pd["freight"] += freight_share

    if unmapped:
        print(f"\nUnmapped products (top 20):")
        for name, count in sorted(unmapped.items(), key=lambda x: -x[1])[:20]:
            print(f"  {name}: {count}")

    print(f"Spare parts skipped: {spare_parts_skipped}")
    print(f"Skipped statuses: {skipped_statuses}")

    return dict(product_data)


# ── Step 4: Push to Google Sheets (categorized layout) ──────────────────────
def make_product_row(product, data, r):
    revenue = round(data["revenue"], 2)
    freight = round(data["freight"], 2)
    return [
        product, revenue, f"=I{r}+K{r}", f"=B{r}-C{r}",
        f'=IF(B{r}=0,"",D{r}/B{r})', f'=IF(J{r}=0,"",D{r}/J{r})',
        data["total_orders"], data["shipped"], f"=R{r}*H{r}",
        data["delivered"], freight, data["rto"], data["in_transit"],
        f'=IF(H{r}=0,"",L{r}/H{r})', f'=IF(G{r}=0,"",H{r}/G{r})',
        f'=IF(G{r}=0,"",J{r}/G{r})', f'=IF(G{r}=0,"",(G{r}-H{r})/G{r})',
        COGS_MAP.get(product, 0),
    ]


def make_subtotal_row(label, first, last, r):
    return [
        label, f"=SUM(B{first}:B{last})", f"=I{r}+K{r}", f"=B{r}-C{r}",
        f'=IF(B{r}=0,"",D{r}/B{r})', f'=IF(J{r}=0,"",D{r}/J{r})',
        f"=SUM(G{first}:G{last})", f"=SUM(H{first}:H{last})",
        f"=SUM(I{first}:I{last})", f"=SUM(J{first}:J{last})",
        f"=SUM(K{first}:K{last})", f"=SUM(L{first}:L{last})",
        f"=SUM(M{first}:M{last})",
        f'=IF(H{r}=0,"",L{r}/H{r})', f'=IF(G{r}=0,"",H{r}/G{r})',
        f'=IF(G{r}=0,"",J{r}/G{r})', f'=IF(G{r}=0,"",(G{r}-H{r})/G{r})', "",
    ]


def push_to_gsheet(product_data):
    print("\nPushing December 2025 MIS to Google Sheets...")

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(CREDS_FILE, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_url(SHEET_URL)

    ws_title = "December 2025 MIS"
    try:
        ws = sh.worksheet(ws_title)
        ws.clear()
        print(f"  Cleared existing '{ws_title}' worksheet")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=ws_title, rows=60, cols=18)
        print(f"  Created new '{ws_title}' worksheet")

    # Build categorized layout
    all_rows = [SHEET_HEADERS]
    category_headers = []
    subtotal_rows = []
    row_num = 2
    subtotal_refs = []

    for cat in CATEGORIES:
        # Category header row
        all_rows.append([cat["name"]] + [""] * 17)
        category_headers.append((row_num, cat["color"]))
        row_num += 1

        first_product_row = row_num
        products_in_cat = 0

        for product in cat["products"]:
            data = product_data.get(product)
            if not data or data["total_orders"] == 0:
                continue
            all_rows.append(make_product_row(product, data, row_num))
            products_in_cat += 1
            row_num += 1

        if products_in_cat > 0:
            last_product_row = row_num - 1
            all_rows.append(make_subtotal_row(
                f"{cat['name']} — Subtotal", first_product_row, last_product_row, row_num))
            subtotal_rows.append((row_num, cat["color"]))
            subtotal_refs.append(row_num)
            row_num += 1
        else:
            all_rows.append(["(no orders this month)"] + [""] * 17)
            row_num += 1

        # Spacer
        all_rows.append([""] * 18)
        row_num += 1

    # Grand Total
    t = row_num
    grand_total = ["GRAND TOTAL"]
    for col_idx in range(1, 18):
        col_letter = chr(ord("A") + col_idx)
        if col_letter == "C":
            grand_total.append(f"=I{t}+K{t}")
        elif col_letter == "D":
            grand_total.append(f"=B{t}-C{t}")
        elif col_letter == "E":
            grand_total.append(f'=IF(B{t}=0,"",D{t}/B{t})')
        elif col_letter == "F":
            grand_total.append(f'=IF(J{t}=0,"",D{t}/J{t})')
        elif col_letter == "N":
            grand_total.append(f'=IF(H{t}=0,"",L{t}/H{t})')
        elif col_letter == "O":
            grand_total.append(f'=IF(G{t}=0,"",H{t}/G{t})')
        elif col_letter == "P":
            grand_total.append(f'=IF(G{t}=0,"",J{t}/G{t})')
        elif col_letter == "Q":
            grand_total.append(f'=IF(G{t}=0,"",(G{t}-H{t})/G{t})')
        elif col_letter == "R":
            grand_total.append("")
        elif col_letter in ("B", "G", "H", "I", "J", "K", "L", "M"):
            refs = "+".join(f"{col_letter}{sr}" for sr in subtotal_refs)
            grand_total.append(f"={refs}")
        else:
            grand_total.append("")
    all_rows.append(grand_total)

    # Push data
    ws.update(range_name="A1", values=all_rows, value_input_option="USER_ENTERED")
    print(f"  Written {len(all_rows)} rows")

    # ── Formatting ──
    ws.format("A1:R1", {
        "backgroundColor": {"red": 0.157, "green": 0.255, "blue": 0.459},
        "textFormat": {"bold": True, "fontSize": 11,
                       "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        "horizontalAlignment": "CENTER",
    })
    time.sleep(1)

    for rn, color in category_headers:
        ws.format(f"A{rn}:R{rn}", {
            "backgroundColor": color,
            "textFormat": {"bold": True, "fontSize": 11,
                           "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        })
        time.sleep(0.5)

    for rn, color in subtotal_rows:
        light = {k: min(1, v * 0.6 + 0.4) for k, v in color.items()}
        ws.format(f"A{rn}:R{rn}", {
            "backgroundColor": light,
            "textFormat": {"bold": True},
        })
        time.sleep(0.5)

    ws.format(f"A{t}:R{t}", {
        "backgroundColor": {"red": 0.20, "green": 0.20, "blue": 0.20},
        "textFormat": {"bold": True, "fontSize": 11,
                       "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
    })
    time.sleep(0.5)

    for col in ["B", "C", "D", "F", "I", "K"]:
        ws.format(f"{col}2:{col}{t}", {"numberFormat": {"type": "NUMBER", "pattern": "₹#,##0"}})
    time.sleep(1)

    for col in ["E", "N", "O", "P", "Q"]:
        ws.format(f"{col}2:{col}{t}", {"numberFormat": {"type": "PERCENT", "pattern": "0.0%"}})
    time.sleep(0.5)

    ws.freeze(rows=1, cols=1)
    print(f"  Done! Sheet: {SHEET_URL}")


# ── Main ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("  DECEMBER 2025 MIS GENERATOR")
    print("=" * 60)

    # Step 1: Fetch orders (or load from cache)
    raw_path = os.path.join(BASE_DIR, "dec_orders_raw.json")
    if os.path.exists(raw_path):
        print("Loading cached December orders from dec_orders_raw.json...")
        with open(raw_path) as f:
            orders = json.load(f)
        print(f"Loaded {len(orders)} orders from cache")
    else:
        orders = fetch_december_orders()
        if orders is None:
            print("Failed to fetch orders. Exiting.")
            exit(1)
        with open(raw_path, "w") as f:
            json.dump(orders, f, default=str)
        print(f"Raw orders saved to: {raw_path}")

    # Step 2: Load freight
    freight_map = load_freight_data()

    # Step 3: Process orders
    product_data = process_orders(orders, freight_map)

    # Save processed data
    processed_path = os.path.join(BASE_DIR, "dec_mis_data.json")
    with open(processed_path, "w") as f:
        json.dump(product_data, f, indent=2, default=str)
    print(f"Processed data saved to: {processed_path}")

    # Step 4: Push to Google Sheets
    push_to_gsheet(product_data)

    # Summary
    print("\n" + "=" * 60)
    print("  DECEMBER 2025 MIS SUMMARY")
    print("=" * 60)
    total_orders = sum(d["total_orders"] for d in product_data.values())
    total_delivered = sum(d["delivered"] for d in product_data.values())
    total_revenue = sum(d["revenue"] for d in product_data.values())
    total_rto = sum(d["rto"] for d in product_data.values())
    total_shipped = sum(d["shipped"] for d in product_data.values())
    total_freight = sum(d["freight"] for d in product_data.values())

    print(f"  Total Orders:     {total_orders:,}")
    print(f"  Total Shipped:    {total_shipped:,}")
    print(f"  Total Delivered:  {total_delivered:,}")
    print(f"  Total RTO:        {total_rto:,}")
    print(f"  Total Revenue:    ₹{total_revenue:,.0f}")
    print(f"  Total Freight:    ₹{total_freight:,.0f}")
    if total_shipped > 0:
        print(f"  RTO%:             {total_rto/total_shipped*100:.1f}%")
    if total_orders > 0:
        print(f"  Delivered%:       {total_delivered/total_orders*100:.1f}%")
    print("=" * 60)
